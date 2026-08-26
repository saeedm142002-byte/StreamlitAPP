import streamlit as st
import pandas as pd
import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from io import BytesIO
import torch
import torch.nn.functional as F
from datetime import datetime, date, time, timedelta
from openpyxl.worksheet.table import Table, TableStyleInfo

import plotly.express as px
import plotly.express as px
import textwrap

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ======================
# MODEL
# ======================
from pathlib import Path
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import streamlit as st

repo_id = "Saeed1233/AraBERTSNB"

@st.cache_resource
def load_model():
    tokenizer = AutoTokenizer.from_pretrained(repo_id)
    model = AutoModelForSequenceClassification.from_pretrained(repo_id)
    model.eval()
    return tokenizer, model

tokenizer, model = load_model()


import torch

def predict_text(text):
    inputs = tokenizer(
        str(text),
        return_tensors="pt",
        truncation=True,
        padding=True,
        max_length=512
    )

    with torch.no_grad():
        outputs = model(**inputs)

    probs = torch.softmax(outputs.logits, dim=1)

    pred = torch.argmax(probs, dim=1).item()   # 0 أو 1
    confidence = probs[0][pred].item() * 100

    return pred, round(confidence, 2)


# ======================
# PAGE CONFIG
# ======================
st.set_page_config(
    page_title="Saeed Mohamed",
    page_icon="❤️🦅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ======================
# SESSION STATE
# ======================
if "page" not in st.session_state:
    st.session_state.page = "الوعود القائمة و المكسورة"

if "sub_page" not in st.session_state:
    st.session_state.sub_page = "اهمال"


# ======================
# SIDEBAR
# ======================
pages = [
    ("الوعود القائمة و المكسورة", "📊"),
    ("الاهمال", "⚠️"),
    ("التوزيع", "📈"),
    ("النشاط", "⚡"),
    ("اخطاء الحالات", "❌"),
    ("التدوير", "🔄")
    
]

with st.sidebar:
    st.markdown("### ❤️🦅 لوحة التحكم")

    for page_name, icon in pages:
        if st.button(f"{icon} {page_name}", use_container_width=True):
            st.session_state.page = page_name
            st.rerun()

    # sub menu فقط للاهمال
    if st.session_state.page == "الاهمال":
        st.markdown("---")
        st.markdown("### خيارات الإهمال")

        for sub in ["اهمال", "متابعة اهمال"]:
            if st.button(sub, use_container_width=True):
                st.session_state.sub_page = sub
                st.rerun()


# ======================
# MAIN ROUTER
# ======================
page = st.session_state.page


# ======================
# PAGE 1 - الوعود القائمة
# ======================
if page == "الوعود القائمة و المكسورة":
    import pandas as pd
    import plotly.express as px
    from io import BytesIO
    import traceback

    # ============================================================
    # 🎨 نظام تصميم مودرن شامل (CSS)
    # ============================================================
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800&display=swap');

        html, body, [class*="css"] {
            font-family: 'Tajawal', sans-serif;
        }

        /* ===== خلفية عامة هادئة ===== */
        .main .block-container {
            padding-top: 1.2rem;
            padding-bottom: 3rem;
        }

        /* ===== الهيدر الرئيسي ===== */
        .promises-header {
            position: relative;
            overflow: hidden;
            background: linear-gradient(120deg, #0a3d2c 0%, #00693E 50%, #0f7a4a 100%);
            padding: 32px 34px;
            border-radius: 20px;
            margin-bottom: 26px;
            box-shadow: 0 10px 30px rgba(0,105,62,0.28);
        }
        .promises-header::after {
            content: "";
            position: absolute;
            top: -60px; left: -40px;
            width: 220px; height: 220px;
            background: radial-gradient(circle, rgba(255,255,255,0.10) 0%, rgba(255,255,255,0) 70%);
            border-radius: 50%;
        }
        .promises-header h1 {
            color: #fff;
            margin: 0;
            font-size: 27px;
            font-weight: 800;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .promises-header p {
            color: #cfe9db;
            margin: 8px 0 0 0;
            font-size: 14.5px;
            font-weight: 500;
        }
        .header-badge {
            display: inline-block;
            background: rgba(255,255,255,0.14);
            color: #fff;
            font-size: 12px;
            font-weight: 700;
            padding: 4px 12px;
            border-radius: 999px;
            margin-top: 12px;
            border: 1px solid rgba(255,255,255,0.25);
        }

        /* ===== بطاقات KPI ===== */
        .kpi-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
            margin-bottom: 6px;
        }
        .kpi-card {
            position: relative;
            background: #ffffff;
            border-radius: 18px;
            padding: 20px 22px;
            box-shadow: 0 6px 20px rgba(17,24,39,0.07);
            border: 1px solid #eef1ef;
            overflow: hidden;
        }
        .kpi-card::before {
            content: "";
            position: absolute;
            top: 0; right: 0;
            width: 90px; height: 90px;
            background: radial-gradient(circle, rgba(0,105,62,0.08) 0%, rgba(0,105,62,0) 70%);
        }
        .kpi-card.broken::before {
            background: radial-gradient(circle, rgba(163,58,58,0.10) 0%, rgba(163,58,58,0) 70%);
        }
        .kpi-top {
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .kpi-icon {
            width: 40px; height: 40px;
            border-radius: 12px;
            display: flex; align-items: center; justify-content: center;
            font-size: 19px;
            background: #e7f4ec;
        }
        .kpi-card.broken .kpi-icon { background: #f8e8e7; }
        .kpi-label {
            font-size: 13px;
            color: #6b7280;
            font-weight: 700;
            margin-top: 12px;
        }
        .kpi-value {
            font-size: 32px;
            font-weight: 800;
            color: #0f172a;
            margin-top: 2px;
        }
        .kpi-sub {
            font-size: 12px;
            color: #9aa4b2;
            font-weight: 600;
            margin-top: 4px;
        }

        /* ===== عناوين الأقسام ===== */
        .section-title {
            display: flex;
            align-items: center;
            gap: 10px;
            background: linear-gradient(90deg, #f3f8f5 0%, #ffffff 100%);
            border-radius: 12px;
            padding: 12px 18px;
            margin: 26px 0 14px 0;
            border-right: 5px solid #00693E;
            font-weight: 800;
            font-size: 16.5px;
            color: #0f3d2e;
        }

        /* ===== لوحة الفلاتر ===== */
        .filter-panel {
            background: #f9fbfa;
            border: 1px solid #e6ede9;
            border-radius: 16px;
            padding: 16px 18px 4px 18px;
            margin-bottom: 18px;
        }
        .filter-panel-title {
            font-size: 13.5px;
            font-weight: 800;
            color: #374151;
            margin-bottom: 8px;
            display: flex; align-items: center; gap: 6px;
        }

        /* ===== بطاقة تحيط بالشارت ===== */
        .chart-card {
            background: #ffffff;
            border-radius: 18px;
            padding: 14px 16px 4px 16px;
            border: 1px solid #eef1ef;
            box-shadow: 0 4px 14px rgba(17,24,39,0.05);
            margin-bottom: 18px;
        }
        .chart-card-title {
            font-weight: 800;
            font-size: 14.5px;
            color: #0f172a;
            margin-bottom: 4px;
        }

        /* ===== تبويبات (Tabs) ===== */
        button[data-baseweb="tab"] {
            font-weight: 700 !important;
            font-size: 14.5px !important;
        }
        div[data-baseweb="tab-list"] {
            gap: 6px;
        }
        button[aria-selected="true"] {
            background: #eaf5ef !important;
            color: #00693E !important;
            border-radius: 10px 10px 0 0 !important;
        }

        /* ===== رفع الملف ===== */
        div[data-testid="stFileUploader"] {
            border: 2px dashed #00693E44;
            border-radius: 16px;
            padding: 8px;
            background: #fafffc;
        }

        /* ===== أزرار التحميل ===== */
        div[data-testid="stDownloadButton"] button {
            border-radius: 12px !important;
            font-weight: 700 !important;
            border: 1px solid #d8e6de !important;
            transition: all 0.15s ease-in-out;
        }
        div[data-testid="stDownloadButton"] button:hover {
            border-color: #00693E !important;
            color: #00693E !important;
            transform: translateY(-1px);
        }

        /* ===== حالة فارغة ===== */
        .empty-state {
            text-align: center;
            padding: 26px 10px;
            color: #9aa4b2;
            font-weight: 600;
            font-size: 14px;
        }

        /* ===== صندوق الأخطاء ===== */
        .error-box {
            background: #fdecea;
            border: 1px solid #f5c2c0;
            border-radius: 14px;
            padding: 18px 20px;
            color: #7a1f1a;
            font-weight: 600;
            line-height: 2;
        }
        .error-box code {
            background: #fbe0de;
            padding: 2px 7px;
            border-radius: 6px;
            font-weight: 700;
        }
        .error-title {
            font-size: 16px;
            font-weight: 800;
            margin-bottom: 6px;
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="promises-header">
        <h1>📊 الوعود القائمة / الوعود المكسورة</h1>
        <p>تتبع الوعود بالسداد، معرفة المكسور منها والقائم، وتحليل الأداء لكل مشرف وموظف</p>
        <span class="header-badge">تحديث لحظي عند رفع الملف</span>
    </div>
    """, unsafe_allow_html=True)

    # ============================================================
    # 🛠️ أدوات مساعدة للتحقق من الأعمدة وعرض الأخطاء بدقة
    # ============================================================
    def require_columns(df, required_cols, step_name):
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise KeyError("STEP::" + step_name + "::MISSING::" + "|".join(missing))

    def show_error(exc):
        msg = str(exc)
        if msg.startswith("STEP::"):
            try:
                _, step_name, _, cols_part = msg.split("::", 3)
                cols_list = cols_part.split("|")
                cols_html = "".join([f"<li><code>{c}</code></li>" for c in cols_list])
                st.markdown(f"""
                <div class="error-box">
                    <div class="error-title">❌ حصل خطأ أثناء تنفيذ خطوة: <code>{step_name}</code></div>
                    الأعمدة التالية غير موجودة في الملف المرفوع:
                    <ul>{cols_html}</ul>
                    تأكد إن أسماء الأعمدة في ملف الإكسيل مطابقة تمامًا (بما فيها المسافات والحروف).
                </div>
                """, unsafe_allow_html=True)
            except Exception:
                st.error(f"❌ خطأ غير متوقع: {msg}")
        else:
            st.markdown(f"""
            <div class="error-box">
                <div class="error-title">❌ حصل خطأ غير متوقع</div>
                <b>نوع الخطأ:</b> <code>{type(exc).__name__}</code><br>
                <b>تفاصيل:</b> {msg}
            </div>
            """, unsafe_allow_html=True)
            with st.expander("🔍 تفاصيل تقنية (Traceback)"):
                st.code(traceback.format_exc())

    # ==========================================
    # اختيار المسار: NPL&Dpd60 أو SNB
    # ==========================================
    portfolio_type = st.radio(
        "اختار نوع المحفظة",
        options=["NPL&Dpd60", "SNB"],
        horizontal=True
    )

    portfolio_file = st.file_uploader(
        "رفع ملف المحفظة",
        type=["xlsx", "xls"]
    )

    # ============================================================
    # الدالة المعالجة الأساسية - Cached
    # ============================================================
    @st.cache_data(show_spinner="جاري معالجة الملف...")
    def process_portfolio(file_bytes, portfolio_type_):
        try:
            df = pd.read_excel(BytesIO(file_bytes))
        except Exception as e:
            raise KeyError(f"STEP::قراءة ملف الإكسيل::MISSING::{e}")

        df = df.iloc[1:].reset_index(drop=True)

        base_required = [
            "Sales Team", "Salesperson", "Final State", "Sub State",
            "Follow up Due Date", "Follow up Last Date"
        ]
        require_columns(df, base_required, "التحقق من الأعمدة الأساسية")

        if portfolio_type_ == "NPL&Dpd60":
            require_columns(
                df,
                ["حالة المعالجة - التمويل", "ملاحظات-التمويل"],
                "التحقق من أعمدة مسار NPL&Dpd60"
            )

        text_cols = [
            "Sales Team", "Salesperson", "Final State", "Sub State",
            "حالة المعالجة - التمويل", "ملاحظات-التمويل"
        ]
        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace(["nan", "None", ""], pd.NA)

        df["Follow up Due Date"] = pd.to_datetime(df["Follow up Due Date"], errors="coerce").dt.normalize()
        df["Follow up Last Date"] = pd.to_datetime(df["Follow up Last Date"], errors="coerce").dt.normalize()

        today = pd.Timestamp.today().normalize()

        if portfolio_type_ == "NPL&Dpd60":
            base = df.copy()
            base = base[base["Sales Team"] != "Sara || Op"]
            base = base[base["Final State"].str.contains("واعد بالسداد", na=False)]
        else:
            base = df.copy()
            allowed_sales_teams = ["SNB II Alsarhan II Naser", "SNB II Alsarhan II Tariq"]
            base = base[base["Sales Team"].isin(allowed_sales_teams)]
            excluded_salespersons = [
                "Closed payments II Alaa SNB", "Hold Companies II SNB2",
                "Abdullah Alsarhan", "Archive Companies II Alaa SNB"
            ]
            base = base[
                (~base["Salesperson"].isin(excluded_salespersons))
                & (base["Salesperson"].notna())
                & (base["Salesperson"].str.strip() != "")
                & (base["Salesperson"].str.lower() != "nan")
            ]
            base = base[base["Sub State"].str.contains("واعد بالسداد", na=False)]

        current = base.copy()
        current = current[current["Follow up Due Date"] == today]
        current = current[current["Follow up Last Date"].notna()]
        current = current[current["Follow up Last Date"] != today]

        broken = base.copy()
        broken = broken[broken["Follow up Due Date"] < today]
        broken = broken[broken["Follow up Last Date"].notna()]
        broken["فرق الايام"] = (broken["Follow up Last Date"] - broken["Follow up Due Date"]).dt.days
        broken = broken[broken["فرق الايام"] < 0]
        broken = broken.drop(columns=["فرق الايام"])

        insert_position = broken.columns.get_loc("Follow up Last Date") + 1
        broken.insert(
            insert_position, "عدد ايام ترحيل الوعد",
            (today - broken["Follow up Due Date"]).dt.days
        )
        broken = broken[broken["عدد ايام ترحيل الوعد"] > 0]

        return current.reset_index(drop=True), broken.reset_index(drop=True)

    if portfolio_file:
        try:
            file_bytes = portfolio_file.getvalue()
            current, broken = process_portfolio(file_bytes, portfolio_type)

            if portfolio_type == "NPL&Dpd60":
                require_columns(current, ["حالة المعالجة - التمويل"], "فلتر حالة المعالجة - التمويل")

                all_status = pd.concat([
                    current["حالة المعالجة - التمويل"], broken["حالة المعالجة - التمويل"]
                ]).dropna().unique().tolist()

                has_empty = (
                    current["حالة المعالجة - التمويل"].isna().any()
                    or broken["حالة المعالجة - التمويل"].isna().any()
                )

                options = sorted([str(x) for x in all_status if str(x).strip() != ""])
                if has_empty:
                    options = ["(فارغ / غير محدد)"] + options

                with st.container():
                    st.markdown('<div class="filter-panel"><div class="filter-panel-title">🔎 فلترة إضافية</div>', unsafe_allow_html=True)
                    selected_status = st.multiselect(
                        "فلتر حسب حالة المعالجة - التمويل (اختياري)",
                        options=options, default=[],
                        help="اختار الحالة اللي عايزها. لو سبتها فاضية هيعرض كل الحالات."
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                if selected_status:
                    def apply_status_filter(df):
                        if "(فارغ / غير محدد)" in selected_status:
                            mask = df["حالة المعالجة - التمويل"].isna()
                            other_selected = [s for s in selected_status if s != "(فارغ / غير محدد)"]
                            if other_selected:
                                mask = mask | df["حالة المعالجة - التمويل"].isin(other_selected)
                            return df[mask]
                        else:
                            return df[df["حالة المعالجة - التمويل"].isin(selected_status)]

                    current = apply_status_filter(current)
                    broken = apply_status_filter(broken)

            # ==========================================
            # 🔢 بطاقات KPI
            # ==========================================
            st.markdown(f"""
            <div class="kpi-grid">
                <div class="kpi-card">
                    <div class="kpi-top">
                        <div class="kpi-icon">✅</div>
                    </div>
                    <div class="kpi-label">الوعود القائمة</div>
                    <div class="kpi-value">{len(current):,}</div>
                    <div class="kpi-sub">حسابات يجب متابعتها اليوم</div>
                </div>
                <div class="kpi-card broken">
                    <div class="kpi-top">
                        <div class="kpi-icon">⚠️</div>
                    </div>
                    <div class="kpi-label">الوعود المكسورة</div>
                    <div class="kpi-value">{len(broken):,}</div>
                    <div class="kpi-sub">حسابات موعد متابعتها عدي</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ==========================================
            # ملفات التحميل (Excel)
            # ==========================================
            output_current = BytesIO()
            with pd.ExcelWriter(output_current, engine="openpyxl") as writer:
                current.to_excel(writer, index=False)
            output_current.seek(0)

            output_broken = BytesIO()
            with pd.ExcelWriter(output_broken, engine="openpyxl") as writer:
                broken.to_excel(writer, index=False)
            output_broken.seek(0)

            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    "📥 تحميل الوعود القائمة", data=output_current,
                    file_name="الوعود_القائمة.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with dl_col2:
                st.download_button(
                    "📥 تحميل الوعود المكسورة", data=output_broken,
                    file_name="الوعود_المكسورة.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            # ============================================================
            # داشبورد الوعود
            # ============================================================
            st.markdown('<div class="section-title">📊 داشبورد الوعود</div>', unsafe_allow_html=True)

            supervisor_col = "ملاحظات-التمويل" if portfolio_type == "NPL&Dpd60" else "Sales Team"
            required_report_cols = [supervisor_col, "Salesperson", "Net Amount", "Account Number"]
            require_columns(current, required_report_cols, "بناء قسم الداشبورد")

            all_supervisors = sorted(
                pd.concat([current[supervisor_col], broken[supervisor_col]]).dropna().unique().tolist()
            )
            all_salespersons = sorted(
                pd.concat([current["Salesperson"], broken["Salesperson"]]).dropna().unique().tolist()
            )

            st.markdown('<div class="filter-panel"><div class="filter-panel-title">🎛️ فلاتر الداشبورد</div>', unsafe_allow_html=True)
            filter_col1, filter_col2 = st.columns(2)
            with filter_col1:
                selected_supervisors = st.multiselect(
                    "فلترة حسب المشرف", options=all_supervisors, default=[],
                    key="promises_supervisor_filter"
                )
            with filter_col2:
                selected_salespersons = st.multiselect(
                    "فلترة حسب الموظف (Salesperson)", options=all_salespersons, default=[],
                    key="promises_salesperson_filter"
                )
            st.markdown('</div>', unsafe_allow_html=True)

            def apply_dashboard_filters(d):
                out = d
                if selected_supervisors:
                    out = out[out[supervisor_col].isin(selected_supervisors)]
                if selected_salespersons:
                    out = out[out["Salesperson"].isin(selected_salespersons)]
                return out

            current_f = apply_dashboard_filters(current)
            broken_f = apply_dashboard_filters(broken)

            def build_pivot_style(d, extra_col=None, extra_label=None, extra_agg="sum"):
                base_cols = ["مبلغ المديونية", "عدد الحسابات"]
                if extra_col:
                    base_cols.append(extra_label)
                out_cols = base_cols + ["الموظف", "المشرف"]

                if d.empty:
                    return pd.DataFrame(columns=out_cols)

                rows = []
                grand_amount = 0.0
                grand_count = 0
                grand_extra = 0.0

                for supervisor, sup_group in d.groupby(supervisor_col, dropna=False):
                    sup_amount = 0.0
                    sup_count = 0
                    sup_extra = 0.0

                    emp_amount = sup_group.groupby("Salesperson", dropna=False)["Net Amount"].sum()
                    emp_count = sup_group.groupby("Salesperson", dropna=False)["Account Number"].nunique()
                    if extra_col:
                        if extra_agg == "sum":
                            emp_extra = sup_group.groupby("Salesperson", dropna=False)[extra_col].sum()
                        else:
                            emp_extra = sup_group.groupby("Salesperson", dropna=False)[extra_col].mean()

                    for salesperson in emp_amount.sort_values(ascending=False).index:
                        amount_val = float(emp_amount.get(salesperson, 0))
                        count_val = int(emp_count.get(salesperson, 0))
                        row = {
                            "مبلغ المديونية": amount_val, "عدد الحسابات": count_val,
                            "الموظف": salesperson, "المشرف": supervisor,
                        }
                        if extra_col:
                            extra_val = float(emp_extra.get(salesperson, 0))
                            row[extra_label] = extra_val
                            sup_extra += extra_val
                        rows.append(row)
                        sup_amount += amount_val
                        sup_count += count_val

                    total_row = {
                        "مبلغ المديونية": sup_amount, "عدد الحسابات": sup_count,
                        "الموظف": f"{supervisor} إجمالي", "المشرف": supervisor,
                    }
                    if extra_col:
                        total_row[extra_label] = (
                            sup_extra if extra_agg == "sum" else (sup_extra / max(sup_count, 1))
                        )
                    rows.append(total_row)

                    grand_amount += sup_amount
                    grand_count += sup_count
                    grand_extra += sup_extra

                grand_row = {
                    "مبلغ المديونية": grand_amount, "عدد الحسابات": grand_count,
                    "الموظف": "الاجمالي", "المشرف": "",
                }
                if extra_col:
                    grand_row[extra_label] = (
                        grand_extra if extra_agg == "sum" else (grand_extra / max(grand_count, 1))
                    )
                rows.append(grand_row)

                return pd.DataFrame(rows)[out_cols]

            def style_summary(d, extra_label=None):
                fmt = {"مبلغ المديونية": "{:,.0f}", "عدد الحسابات": "{:,.0f}"}
                if extra_label:
                    fmt[extra_label] = "{:,.0f}"

                def highlight_rows(row):
                    is_total = (row["الموظف"] == "الاجمالي") or ("إجمالي" in str(row["الموظف"]))
                    if is_total:
                        return ["font-weight: bold; color: #000000; background-color: #eef2f7"] * len(row)
                    return [""] * len(row)

                return d.style.apply(highlight_rows, axis=1).format(fmt)

            tab_current, tab_broken, tab_days, tab_raw = st.tabs([
                "📄 الوعود القائمة", "📄 الوعود المكسورة", "📄 أيام الترحيل", "🗂️ أصل البيانات"
            ])

            with tab_current:
                current_summary = build_pivot_style(current_f)
                if current_summary.empty:
                    st.markdown('<div class="empty-state">لا توجد بيانات مطابقة لهذا الفلتر</div>', unsafe_allow_html=True)
                else:
                    st.dataframe(style_summary(current_summary), use_container_width=True, hide_index=True)

            with tab_broken:
                broken_summary = build_pivot_style(broken_f)
                if broken_summary.empty:
                    st.markdown('<div class="empty-state">لا توجد بيانات مطابقة لهذا الفلتر</div>', unsafe_allow_html=True)
                else:
                    st.dataframe(style_summary(broken_summary), use_container_width=True, hide_index=True)

            with tab_days:
                require_columns(broken_f, ["عدد ايام ترحيل الوعد"], "تقرير عدد أيام ترحيل الوعد")
                days_summary = build_pivot_style(
                    broken_f, extra_col="عدد ايام ترحيل الوعد",
                    extra_label="عدد ايام ترحيل الوعد", extra_agg="sum"
                )
                if days_summary.empty:
                    st.markdown('<div class="empty-state">لا توجد بيانات مطابقة لهذا الفلتر</div>', unsafe_allow_html=True)
                else:
                    st.dataframe(
                        style_summary(days_summary, extra_label="عدد ايام ترحيل الوعد"),
                        use_container_width=True, hide_index=True
                    )

            with tab_raw:
                st.markdown("#### 🗂️ الوعود القائمة")
                st.dataframe(current_f, use_container_width=True, hide_index=True)
                st.markdown("#### 🗂️ الوعود المكسورة")
                st.dataframe(broken_f, use_container_width=True, hide_index=True)

            # ---------------------------
            # الشارتات
            # ---------------------------
            st.markdown('<div class="section-title">📈 الرسوم البيانية</div>', unsafe_allow_html=True)

            SNB_GREEN = "#00693E"
            SNB_GOLD = "#C9A227"
            SNB_RED = "#A33A3A"

            # النص بيبقى أبيض في الوضع الداكن (Dark) وأسود في الوضع الفاتح (Light)
            def get_theme_text_color():
                try:
                    base = st.get_option("theme.base")
                except Exception:
                    base = None
                return "#FFFFFF" if base == "dark" else "#111827"

            AXIS_TEXT_COLOR = get_theme_text_color()
            GRID_COLOR = "#374151" if AXIS_TEXT_COLOR == "#FFFFFF" else "#f1f5f3"
            LABEL_FONT = dict(size=14, family="Tajawal", color=AXIS_TEXT_COLOR)

            def style_fig(fig, angle=-20):
                fig.update_xaxes(tickfont=dict(size=13, family="Tajawal", color=AXIS_TEXT_COLOR))
                fig.update_yaxes(tickfont=dict(size=12, family="Tajawal", color=AXIS_TEXT_COLOR), gridcolor=GRID_COLOR)
                fig.update_layout(
                    height=400, xaxis_tickangle=angle, margin=dict(t=20, b=10, l=10, r=10),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Tajawal"),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                return fig

            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.markdown('<div class="chart-card"><div class="chart-card-title">مبلغ المديونية لكل مشرف (قائمة مقابل مكسورة)</div>', unsafe_allow_html=True)
                sup_amount_current = (
                    current_f.groupby(supervisor_col, dropna=False)["Net Amount"].sum()
                    .reset_index().rename(columns={supervisor_col: "المشرف", "Net Amount": "مبلغ المديونية"})
                )
                sup_amount_current["النوع"] = "قائمة"
                sup_amount_broken = (
                    broken_f.groupby(supervisor_col, dropna=False)["Net Amount"].sum()
                    .reset_index().rename(columns={supervisor_col: "المشرف", "Net Amount": "مبلغ المديونية"})
                )
                sup_amount_broken["النوع"] = "مكسورة"
                sup_amount_combined = pd.concat([sup_amount_current, sup_amount_broken], ignore_index=True)

                if not sup_amount_combined.empty:
                    fig1 = px.bar(
                        sup_amount_combined, x="المشرف", y="مبلغ المديونية", color="النوع",
                        barmode="group", text="مبلغ المديونية",
                        color_discrete_map={"قائمة": SNB_GREEN, "مكسورة": SNB_RED},
                        template="plotly_white"
                    )
                    fig1.update_traces(texttemplate="<b>%{text:,.0f}</b>", textposition="outside", textfont=LABEL_FONT)
                    st.plotly_chart(style_fig(fig1), use_container_width=True)
                else:
                    st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with chart_col2:
                st.markdown('<div class="chart-card"><div class="chart-card-title">متوسط أيام ترحيل الوعد لكل مشرف</div>', unsafe_allow_html=True)
                days_by_sup = (
                    broken_f.groupby(supervisor_col, dropna=False)["عدد ايام ترحيل الوعد"].mean()
                    .reset_index().rename(columns={supervisor_col: "المشرف", "عدد ايام ترحيل الوعد": "متوسط ايام الترحيل"})
                )
                if not days_by_sup.empty:
                    fig3 = px.bar(
                        days_by_sup, x="المشرف", y="متوسط ايام الترحيل", text="متوسط ايام الترحيل",
                        color_discrete_sequence=[SNB_GOLD], template="plotly_white"
                    )
                    fig3.update_traces(texttemplate="<b>%{text:.1f}</b>", textposition="outside", textfont=LABEL_FONT)
                    st.plotly_chart(style_fig(fig3), use_container_width=True)
                else:
                    st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="chart-card"><div class="chart-card-title">أعلى 15 موظف بعدد الحسابات (الوعود المكسورة)</div>', unsafe_allow_html=True)
            emp_count_broken = (
                broken_f.groupby("Salesperson", dropna=False)["Account Number"].nunique()
                .reset_index(name="عدد الحسابات")
                .sort_values("عدد الحسابات", ascending=False).head(15)
            )
            if not emp_count_broken.empty:
                fig2 = px.bar(
                    emp_count_broken, x="Salesperson", y="عدد الحسابات", text="عدد الحسابات",
                    color_discrete_sequence=[SNB_RED], template="plotly_white"
                )
                fig2.update_traces(texttemplate="<b>%{text}</b>", textposition="outside", textfont=LABEL_FONT)
                st.plotly_chart(style_fig(fig2, angle=-30), use_container_width=True)
            else:
                st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        except KeyError as e:
            show_error(e)
        except Exception as e:
            show_error(e)
    else:
        st.markdown('<div class="empty-state">⬆️ ارفع ملف المحفظة عشان يبدأ التحليل</div>', unsafe_allow_html=True)
# ======================
# PAGE 2 - الوعود المكسورة
# ======================
elif page == "الوعود المكسورة":
    st.subheader("🚨 الوعود المكسورة")
    st.write("متابعة الوعود المكسورة")


# ======================
# PAGE 3 - الاهمال
# ======================
elif page == "الاهمال":

    sub = st.session_state.sub_page

    if sub == "اهمال":

        import pandas as pd
        import re
        import traceback
        from io import BytesIO

        # ============================================================
        # 🎨 نظام تصميم مودرن شامل (نفس ستايل صفحة الوعود)
        # ============================================================
        st.markdown("""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800&display=swap');

            html, body, [class*="css"] {
                font-family: 'Tajawal', sans-serif;
            }

            .main .block-container {
                padding-top: 1.2rem;
                padding-bottom: 3rem;
            }

            /* ===== الهيدر الرئيسي ===== */
            .neglect-header {
                position: relative;
                overflow: hidden;
                background: linear-gradient(120deg, #5a1f1a 0%, #A33A3A 50%, #b8493f 100%);
                padding: 32px 34px;
                border-radius: 20px;
                margin-bottom: 26px;
                box-shadow: 0 10px 30px rgba(163,58,58,0.28);
            }
            .neglect-header::after {
                content: "";
                position: absolute;
                top: -60px; left: -40px;
                width: 220px; height: 220px;
                background: radial-gradient(circle, rgba(255,255,255,0.10) 0%, rgba(255,255,255,0) 70%);
                border-radius: 50%;
            }
            .neglect-header h1 {
                color: #fff;
                margin: 0;
                font-size: 27px;
                font-weight: 800;
                display: flex;
                align-items: center;
                gap: 10px;
            }
            .neglect-header p {
                color: #f3dcda;
                margin: 8px 0 0 0;
                font-size: 14.5px;
                font-weight: 500;
            }
            .header-badge {
                display: inline-block;
                background: rgba(255,255,255,0.14);
                color: #fff;
                font-size: 12px;
                font-weight: 700;
                padding: 4px 12px;
                border-radius: 999px;
                margin-top: 12px;
                border: 1px solid rgba(255,255,255,0.25);
            }

            /* ===== بطاقات KPI ===== */
            .kpi-grid {
                display: grid;
                grid-template-columns: 1fr;
                gap: 16px;
                margin-bottom: 6px;
            }
            .kpi-card {
                position: relative;
                background: #ffffff;
                border-radius: 18px;
                padding: 20px 22px;
                box-shadow: 0 6px 20px rgba(17,24,39,0.07);
                border: 1px solid #eef1ef;
                border-left: 6px solid #A33A3A;
                overflow: hidden;
            }
            .kpi-card::before {
                content: "";
                position: absolute;
                top: 0; right: 0;
                width: 90px; height: 90px;
                background: radial-gradient(circle, rgba(163,58,58,0.10) 0%, rgba(163,58,58,0) 70%);
            }
            .kpi-top {
                display: flex;
                align-items: center;
                justify-content: space-between;
            }
            .kpi-icon {
                width: 40px; height: 40px;
                border-radius: 12px;
                display: flex; align-items: center; justify-content: center;
                font-size: 19px;
                background: #f8e8e7;
            }
            .kpi-label {
                font-size: 13px;
                color: #6b7280;
                font-weight: 700;
                margin-top: 12px;
            }
            .kpi-value {
                font-size: 32px;
                font-weight: 800;
                color: #0f172a;
                margin-top: 2px;
            }
            .kpi-sub {
                font-size: 12px;
                color: #9aa4b2;
                font-weight: 600;
                margin-top: 4px;
            }

            /* ===== عناوين الأقسام ===== */
            .section-title {
                display: flex;
                align-items: center;
                gap: 10px;
                background: linear-gradient(90deg, #faf3f2 0%, #ffffff 100%);
                border-radius: 12px;
                padding: 12px 18px;
                margin: 26px 0 14px 0;
                border-right: 5px solid #A33A3A;
                font-weight: 800;
                font-size: 16.5px;
                color: #4a1815;
            }

            /* ===== لوحة الفلاتر ===== */
            .filter-panel {
                background: #f9fbfa;
                border: 1px solid #e6ede9;
                border-radius: 16px;
                padding: 16px 18px 4px 18px;
                margin-bottom: 18px;
            }
            .filter-panel-title {
                font-size: 13.5px;
                font-weight: 800;
                color: #374151;
                margin-bottom: 8px;
                display: flex; align-items: center; gap: 6px;
            }

            /* ===== لوحة اختيار الحالات ===== */
            .states-panel {
                background: #fffaf9;
                border: 1px solid #f1dedd;
                border-radius: 16px;
                padding: 16px 18px;
                margin-bottom: 10px;
            }
            .states-panel-title {
                font-size: 14.5px;
                font-weight: 800;
                color: #4a1815;
                margin-bottom: 8px;
            }

            /* ===== بطاقة تحيط بالشارت ===== */
            .chart-card {
                background: #ffffff;
                border-radius: 18px;
                padding: 14px 16px 4px 16px;
                border: 1px solid #eef1ef;
                box-shadow: 0 4px 14px rgba(17,24,39,0.05);
                margin-bottom: 18px;
            }
            .chart-card-title {
                font-weight: 800;
                font-size: 14.5px;
                color: #0f172a;
                margin-bottom: 4px;
            }

            /* ===== رفع الملف ===== */
            div[data-testid="stFileUploader"] {
                border: 2px dashed #A33A3A44;
                border-radius: 16px;
                padding: 8px;
                background: #fffaf9;
            }

            /* ===== أزرار التحميل ===== */
            div[data-testid="stDownloadButton"] button {
                border-radius: 12px !important;
                font-weight: 700 !important;
                border: 1px solid #ecdcda !important;
                transition: all 0.15s ease-in-out;
            }
            div[data-testid="stDownloadButton"] button:hover {
                border-color: #A33A3A !important;
                color: #A33A3A !important;
                transform: translateY(-1px);
            }

            /* ===== زر تشغيل التقرير ===== */
            div[data-testid="stButton"] button[kind="primary"] {
                background: linear-gradient(120deg, #A33A3A, #8a2f2f) !important;
                border: none !important;
                border-radius: 12px !important;
                font-weight: 800 !important;
                box-shadow: 0 4px 14px rgba(163,58,58,0.25);
            }

            /* ===== حالة فارغة ===== */
            .empty-state {
                text-align: center;
                padding: 26px 10px;
                color: #9aa4b2;
                font-weight: 600;
                font-size: 14px;
            }

            /* ===== صندوق الأخطاء ===== */
            .error-box {
                background: #fdecea;
                border: 1px solid #f5c2c0;
                border-radius: 14px;
                padding: 18px 20px;
                color: #7a1f1a;
                font-weight: 600;
                line-height: 2;
            }
            .error-box code {
                background: #fbe0de;
                padding: 2px 7px;
                border-radius: 6px;
                font-weight: 700;
            }
            .error-title {
                font-size: 16px;
                font-weight: 800;
                margin-bottom: 6px;
            }
        </style>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="neglect-header">
            <h1>⚠️ الاهمال</h1>
            <p>رصد الحسابات المهملة اللي محدّش تابعها لفترة، وتحليل الأداء لكل مشرف وموظف</p>
            <span class="header-badge">تحديث لحظي عند رفع الملف</span>
        </div>
        """, unsafe_allow_html=True)

        # ============================================================
        # 🛠️ أدوات مساعدة للتحقق من الأعمدة وعرض الأخطاء بدقة
        # ============================================================
        def require_columns(df, required_cols, step_name):
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                raise KeyError("STEP::" + step_name + "::MISSING::" + "|".join(missing))

        def show_error(exc):
            msg = str(exc)
            if msg.startswith("STEP::"):
                try:
                    _, step_name, _, cols_part = msg.split("::", 3)
                    cols_list = cols_part.split("|")
                    cols_html = "".join([f"<li><code>{c}</code></li>" for c in cols_list])
                    st.markdown(f"""
                    <div class="error-box">
                        <div class="error-title">❌ حصل خطأ أثناء تنفيذ خطوة: <code>{step_name}</code></div>
                        الأعمدة التالية غير موجودة في الملف المرفوع:
                        <ul>{cols_html}</ul>
                        تأكد إن أسماء الأعمدة في ملف الإكسيل مطابقة تمامًا (بما فيها المسافات والحروف).
                    </div>
                    """, unsafe_allow_html=True)
                except Exception:
                    st.error(f"❌ خطأ غير متوقع: {msg}")
            else:
                st.markdown(f"""
                <div class="error-box">
                    <div class="error-title">❌ حصل خطأ غير متوقع</div>
                    <b>نوع الخطأ:</b> <code>{type(exc).__name__}</code><br>
                    <b>تفاصيل:</b> {msg}
                </div>
                """, unsafe_allow_html=True)
                with st.expander("🔍 تفاصيل تقنية (Traceback)"):
                    st.code(traceback.format_exc())

        # ==========================================
        # اختيار المسار: NPL&Dpd60 أو SNB
        # ==========================================
        neglect_portfolio_type = st.radio(
            "اختار نوع المحفظة",
            options=["NPL&Dpd60", "SNB"],
            horizontal=True,
            key="neglect_portfolio_type"
        )

        # ==========================================
        # سؤال عن عدد الأيام حسب المسار
        # ==========================================
        if neglect_portfolio_type == "NPL&Dpd60":
            neglect_days_threshold = st.number_input(
                "عدد أيام الإهمال (الحد الأدنى)",
                min_value=1, value=3, step=1, key="neglect_npl_days"
            )
        else:
            neglect_days_threshold = st.number_input(
                "فرق عدد ايام من اخر متابعة (الحد الأدنى)",
                min_value=1, value=7, step=1, key="neglect_snb_days"
            )

        uploaded_file = st.file_uploader(
            "رفع ملف المحفظة",
            type=["xlsx", "xls"],
            key="neglect_file_uploader"
        )

        # ==========================================
        # الحالات الافتراضية (default) لكل مسار
        # ==========================================
        NPL_DEFAULT_STATES = [
            "⁠وعد بسداد مبلغ المعالجة",
            "*رافض السداد",
            "*مسجون",
            "*مماطل",
            "تم سداد جزء وليس مبلغ المعالجة",
            "لا يجيب",
            "وعد بسداد تسوية",
            "وعد بسداد جزء من المتأخرات",
            "وعد بسداد قسط",
            "وعد بسداد كامل المتأخرات",
            "وعد بسداد كامل المديونية",
            "يرغب بجدولة المديونية"
        ]

        SNB_DEFAULT_STATES = [
            "تم ابلاغ العميل - اتصال",
            "جدولة",
            "سداد جزئي",
            "قيد التفاوض مع الورثة",
            "مماطل",
            "واعد بالسداد"
        ]

        def normalize_text(x):
            if pd.isna(x):
                return ""
            x = str(x)
            x = re.sub(r"[\u200b\u200c\u200d\u200e\u200f\u2060\ufeff]", "", x)
            x = re.sub(r"\s+", "", x)
            return x.strip()

        # ============================================================
        # الدالة المعالجة الأساسية - Cached
        # ============================================================
        @st.cache_data(show_spinner="جاري معالجة الملف...")
        def process_neglect(file_bytes, portfolio_type_, days_threshold_, selected_states_):

            selected_states_norm = {normalize_text(s) for s in selected_states_}

            try:
                df = pd.read_excel(BytesIO(file_bytes))
            except Exception as e:
                raise KeyError(f"STEP::قراءة ملف الإكسيل::MISSING::{e}")

            df = df.iloc[1:].reset_index(drop=True)

            today = pd.Timestamp.today().normalize()

            if portfolio_type_ == "NPL&Dpd60":

                require_columns(
                    df,
                    ["Sales Team", "Sub State", "Follow up Last Date", "Payment"],
                    "معالجة مسار NPL&Dpd60"
                )

                text_cols = ["Sales Team", "Sub State", "حالة المعالجة - التمويل"]
                for col in text_cols:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip()

                df["Follow up Last Date"] = pd.to_datetime(
                    df["Follow up Last Date"], errors="coerce"
                ).dt.normalize()

                df = df[df["Sales Team"] != "Sara || Op"]

                df["Payment"] = (
                    df["Payment"].astype(str).str.replace(",", "", regex=False).str.strip()
                )
                df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce")
                df = df[df["Payment"] <= 0]

                df["عدد أيام الإهمال"] = (today - df["Follow up Last Date"]).dt.days

                if "فرق عدد ايام من اخر متابعة" in df.columns:
                    df.drop(columns=["فرق عدد ايام من اخر متابعة"], inplace=True)

                insert_position = df.columns.get_loc("Follow up Last Date") + 1
                df.insert(
                    insert_position, "فرق عدد ايام من اخر متابعة",
                    (today - df["Follow up Last Date"]).dt.days
                )

                df = df[df["عدد أيام الإهمال"] >= days_threshold_]
                df = df[df["Sub State"].apply(normalize_text).isin(selected_states_norm)]

                supervisor_col_ = "ملاحظات-التمويل"
                require_columns(df, [supervisor_col_], "التحقق من عمود المشرف (NPL&Dpd60)")

            else:

                require_columns(
                    df,
                    ["Sales Team", "Salesperson", "Sub State", "Follow up Last Date", "Payment"],
                    "معالجة مسار SNB"
                )

                text_cols = ["Sales Team", "Salesperson", "Sub State"]
                for col in text_cols:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip()

                df["Follow up Last Date"] = pd.to_datetime(
                    df["Follow up Last Date"], errors="coerce"
                ).dt.normalize()

                allowed_sales_teams = [
                    "SNB II Alsarhan II Naser",
                    "SNB II Alsarhan II Tariq"
                ]
                df = df[df["Sales Team"].isin(allowed_sales_teams)]

                excluded_salespersons = [
                    "Closed payments II Alaa SNB",
                    "Hold Companies II SNB2",
                    "Abdullah Alsarhan",
                    "Archive Companies II Alaa SNB"
                ]
                df = df[
                    (~df["Salesperson"].isin(excluded_salespersons))
                    & (df["Salesperson"].notna())
                    & (df["Salesperson"].str.strip() != "")
                    & (df["Salesperson"].str.lower() != "nan")
                ]

                df["Payment"] = (
                    df["Payment"].astype(str).str.replace(",", "", regex=False).str.strip()
                )
                df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce")
                df = df[df["Payment"] == 0]

                df = df[
                    df["Sub State"].apply(
                        lambda v: any(t in normalize_text(v) for t in selected_states_norm)
                    )
                ]

                if "فرق عدد ايام من اخر متابعة" in df.columns:
                    df.drop(columns=["فرق عدد ايام من اخر متابعة"], inplace=True)

                insert_position = df.columns.get_loc("Follow up Last Date") + 1
                df.insert(
                    insert_position, "فرق عدد ايام من اخر متابعة",
                    (today - df["Follow up Last Date"]).dt.days
                )

                df = df[df["فرق عدد ايام من اخر متابعة"] >= days_threshold_]

                supervisor_col_ = "Sales Team"

            return df.reset_index(drop=True), supervisor_col_

        # ============================================================
        # قراءة الحالات الموجودة في الملف (بدون تنفيذ الفلترة الكاملة)
        # ============================================================
        @st.cache_data(show_spinner="جاري قراءة الحالات...")
        def read_sub_states(file_bytes_):
            df_raw = pd.read_excel(BytesIO(file_bytes_))
            df_raw = df_raw.iloc[1:].reset_index(drop=True)
            if "Sub State" not in df_raw.columns:
                return []
            vals = df_raw["Sub State"].dropna().astype(str).str.strip()
            vals = vals[vals != ""]
            return sorted(vals.unique().tolist())

        if uploaded_file:
          try:
            file_bytes = uploaded_file.getvalue()

            require_columns(
                pd.read_excel(BytesIO(file_bytes), nrows=1),
                ["Sub State"],
                "قراءة الحالات (Sub State) من الملف"
            )

            all_states = read_sub_states(file_bytes)

            pool_key = f"neglect_pools_{neglect_portfolio_type}"
            sig_key = f"{pool_key}_sig"
            run_key = f"neglect_run_{neglect_portfolio_type}"
            file_signature = (uploaded_file.name, len(file_bytes), neglect_portfolio_type)

            # تبني البوكسين أول مرة فقط، أو لو الملف/المسار اتغير
            if st.session_state.get(sig_key) != file_signature:
                default_states = (
                    NPL_DEFAULT_STATES if neglect_portfolio_type == "NPL&Dpd60"
                    else SNB_DEFAULT_STATES
                )
                default_norm = [normalize_text(s) for s in default_states]

                if neglect_portfolio_type == "NPL&Dpd60":
                    neglect_states = [s for s in all_states if normalize_text(s) in default_norm]
                else:
                    neglect_states = [
                        s for s in all_states
                        if any(t in normalize_text(s) for t in default_norm)
                    ]

                other_states = [s for s in all_states if s not in neglect_states]

                st.session_state[pool_key] = {
                    "not_neglect": sorted(other_states),
                    "neglect": sorted(neglect_states)
                }
                st.session_state[sig_key] = file_signature
                st.session_state[run_key] = False

            pools = st.session_state[pool_key]

            st.markdown('<div class="section-title">🗂️ اختيار حالات الإهمال</div>', unsafe_allow_html=True)

            box1, box2 = st.columns(2)

            with box1:
                st.markdown('<div class="states-panel"><div class="states-panel-title">📋 باقي الحالات</div>', unsafe_allow_html=True)
                search_other = st.text_input(
                    "🔍 بحث",
                    key=f"search_other_{neglect_portfolio_type}"
                )
                filtered_other = (
                    [s for s in pools["not_neglect"] if search_other.strip() in s]
                    if search_other.strip() else pools["not_neglect"]
                )

                pick_other = st.selectbox(
                    "اختار حالة",
                    options=filtered_other or ["— لا يوجد —"],
                    key=f"pick_other_{neglect_portfolio_type}"
                )
                if st.button("➡️ Insert", key=f"insert_{neglect_portfolio_type}",
                             disabled=not filtered_other):
                    pools["not_neglect"].remove(pick_other)
                    pools["neglect"].append(pick_other)
                    pools["neglect"].sort()
                    st.session_state[run_key] = False
                    st.rerun()

                st.dataframe(
                    pd.DataFrame({"الحالة": filtered_other}),
                    use_container_width=True, hide_index=True
                )
                st.markdown('</div>', unsafe_allow_html=True)

            with box2:
                st.markdown('<div class="states-panel"><div class="states-panel-title">⚠️ حالات الإهمال</div>', unsafe_allow_html=True)
                search_neglect = st.text_input(
                    "🔍 بحث",
                    key=f"search_neglect_{neglect_portfolio_type}"
                )
                filtered_neglect = (
                    [s for s in pools["neglect"] if search_neglect.strip() in s]
                    if search_neglect.strip() else pools["neglect"]
                )

                pick_neglect = st.selectbox(
                    "اختار حالة",
                    options=filtered_neglect or ["— لا يوجد —"],
                    key=f"pick_neglect_{neglect_portfolio_type}"
                )
                if st.button("🗑️ Delete", key=f"delete_{neglect_portfolio_type}",
                             disabled=not filtered_neglect):
                    pools["neglect"].remove(pick_neglect)
                    pools["not_neglect"].append(pick_neglect)
                    pools["not_neglect"].sort()
                    st.session_state[run_key] = False
                    st.rerun()

                st.dataframe(
                    pd.DataFrame({"الحالة": filtered_neglect}),
                    use_container_width=True, hide_index=True
                )
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            run_clicked = st.button(
                "🚀 عمل الاهمال",
                key=f"run_btn_{neglect_portfolio_type}",
                type="primary",
                disabled=not pools["neglect"]
            )
            if run_clicked:
                st.session_state[run_key] = True

            # ============================================================
            # التقرير + الداشبورد - يظهروا بس بعد "عمل الاهمال"
            # ============================================================
            if st.session_state.get(run_key):

                selected_states = tuple(sorted(pools["neglect"]))
                df, supervisor_col = process_neglect(
                    file_bytes, neglect_portfolio_type, neglect_days_threshold, selected_states
                )

                # ============================================================
                # فلتر "حالة المعالجة - التمويل" (اختياري من المستخدم)
                # ============================================================
                status_col = "حالة المعالجة - التمويل"

                if neglect_portfolio_type == "NPL&Dpd60" and status_col in df.columns:

                    raw_statuses = df[status_col].unique().tolist()

                    status_options = []
                    for v in raw_statuses:
                        if pd.isna(v) or str(v).strip().lower() in ("nan", ""):
                            status_options.append("بدون حالة (فاضي)")
                        else:
                            status_options.append(str(v).strip())
                    status_options = sorted(set(status_options))

                    st.markdown('<div class="filter-panel"><div class="filter-panel-title">🔎 فلترة إضافية</div>', unsafe_allow_html=True)
                    selected_statuses = st.multiselect(
                        "فلترة حسب حالة المعالجة - التمويل",
                        options=status_options,
                        default=[],
                        key=f"neglect_processing_status_filter_{neglect_portfolio_type}",
                        help="سيب الاختيار فاضي عشان تعرض كل الحالات"
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

                    if selected_statuses:

                        def status_mask(series):
                            normalized = series.apply(
                                lambda v: "بدون حالة (فاضي)"
                                if (pd.isna(v) or str(v).strip().lower() in ("nan", ""))
                                else str(v).strip()
                            )
                            return normalized.isin(selected_statuses)

                        df = df[status_mask(df[status_col])].reset_index(drop=True)

                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)

                # ==========================================
                # 🔢 بطاقة KPI
                # ==========================================
                st.markdown(f"""
                <div class="kpi-grid">
                    <div class="kpi-card">
                        <div class="kpi-top">
                            <div class="kpi-icon">⚠️</div>
                        </div>
                        <div class="kpi-label">إجمالي الحسابات المهملة</div>
                        <div class="kpi-value">{len(df):,}</div>
                        <div class="kpi-sub">حسابات مطابقة لحدود الإهمال المحددة</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                st.download_button(
                    "📥 تحميل تقرير الإهمال",
                    data=output,
                    file_name="الاهمال.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                st.markdown('<div class="section-title">📊 داشبورد الإهمال</div>', unsafe_allow_html=True)

                required_report_cols = [supervisor_col, "Salesperson", "Net Amount", "Account Number"]
                require_columns(df, required_report_cols, "بناء قسم الداشبورد")

                import plotly.express as px

                all_supervisors = sorted(df[supervisor_col].dropna().unique().tolist())
                all_salespersons = sorted(df["Salesperson"].dropna().unique().tolist())

                st.markdown('<div class="filter-panel"><div class="filter-panel-title">🎛️ فلاتر الداشبورد</div>', unsafe_allow_html=True)
                filter_col1, filter_col2 = st.columns(2)
                with filter_col1:
                    selected_supervisors = st.multiselect(
                        "فلترة حسب المشرف", options=all_supervisors, default=[],
                        key="neglect_supervisor_filter"
                    )
                with filter_col2:
                    selected_salespersons = st.multiselect(
                        "فلترة حسب الموظف (Salesperson)", options=all_salespersons, default=[],
                        key="neglect_salesperson_filter"
                    )
                st.markdown('</div>', unsafe_allow_html=True)

                def apply_dashboard_filters(d):
                    out = d
                    if selected_supervisors:
                        out = out[out[supervisor_col].isin(selected_supervisors)]
                    if selected_salespersons:
                        out = out[out["Salesperson"].isin(selected_salespersons)]
                    return out

                df_f = apply_dashboard_filters(df)

                def build_pivot_style(d, extra_col=None, extra_label=None, extra_agg="sum"):
                    base_cols = ["مبلغ المديونية", "عدد الحسابات"]
                    if extra_col:
                        base_cols.append(extra_label)
                    out_cols = base_cols + ["الموظف", "المشرف"]

                    if d.empty:
                        return pd.DataFrame(columns=out_cols)

                    rows = []
                    grand_amount = 0.0
                    grand_count = 0
                    grand_extra = 0.0

                    for supervisor, sup_group in d.groupby(supervisor_col, dropna=False):
                        sup_amount = 0.0
                        sup_count = 0
                        sup_extra = 0.0

                        emp_amount = sup_group.groupby("Salesperson", dropna=False)["Net Amount"].sum()
                        emp_count = sup_group.groupby("Salesperson", dropna=False)["Account Number"].nunique()
                        if extra_col:
                            if extra_agg == "sum":
                                emp_extra = sup_group.groupby("Salesperson", dropna=False)[extra_col].sum()
                            else:
                                emp_extra = sup_group.groupby("Salesperson", dropna=False)[extra_col].mean()

                        for salesperson in emp_amount.sort_values(ascending=False).index:
                            amount_val = float(emp_amount.get(salesperson, 0))
                            count_val = int(emp_count.get(salesperson, 0))
                            row = {
                                "مبلغ المديونية": amount_val,
                                "عدد الحسابات": count_val,
                                "الموظف": salesperson,
                                "المشرف": supervisor,
                            }
                            if extra_col:
                                extra_val = float(emp_extra.get(salesperson, 0))
                                row[extra_label] = extra_val
                                sup_extra += extra_val
                            rows.append(row)
                            sup_amount += amount_val
                            sup_count += count_val

                        total_row = {
                            "مبلغ المديونية": sup_amount,
                            "عدد الحسابات": sup_count,
                            "الموظف": f"{supervisor} إجمالي",
                            "المشرف": supervisor,
                        }
                        if extra_col:
                            total_row[extra_label] = (
                                sup_extra if extra_agg == "sum" else (sup_extra / max(sup_count, 1))
                            )
                        rows.append(total_row)

                        grand_amount += sup_amount
                        grand_count += sup_count
                        grand_extra += sup_extra

                    grand_row = {
                        "مبلغ المديونية": grand_amount,
                        "عدد الحسابات": grand_count,
                        "الموظف": "الاجمالي",
                        "المشرف": "",
                    }
                    if extra_col:
                        grand_row[extra_label] = (
                            grand_extra if extra_agg == "sum" else (grand_extra / max(grand_count, 1))
                        )
                    rows.append(grand_row)

                    return pd.DataFrame(rows)[out_cols]

                def style_summary(d, extra_label=None):
                    fmt = {"مبلغ المديونية": "{:,.0f}", "عدد الحسابات": "{:,.0f}"}
                    if extra_label:
                        fmt[extra_label] = "{:,.0f}"

                    def highlight_rows(row):
                        is_total = (row["الموظف"] == "الاجمالي") or ("إجمالي" in str(row["الموظف"]))
                        if is_total:
                            return ["font-weight: bold; color: #000000; background-color: #eef2f7"] * len(row)
                        return [""] * len(row)

                    return d.style.apply(highlight_rows, axis=1).format(fmt)

                day_col_label = (
                    "عدد أيام الإهمال" if neglect_portfolio_type == "NPL&Dpd60"
                    else "فرق عدد ايام من اخر متابعة"
                )

                tab_report, tab_days, tab_raw = st.tabs([
                    "📄 تقرير الإهمال", f"📄 {day_col_label}", "🗂️ أصل البيانات"
                ])

                with tab_report:
                    neglect_summary = build_pivot_style(df_f)
                    if neglect_summary.empty:
                        st.markdown('<div class="empty-state">لا توجد بيانات مطابقة لهذا الفلتر</div>', unsafe_allow_html=True)
                    else:
                        st.dataframe(style_summary(neglect_summary), use_container_width=True, hide_index=True)

                with tab_days:
                    require_columns(df_f, ["فرق عدد ايام من اخر متابعة"], f"تقرير {day_col_label}")
                    days_summary = build_pivot_style(
                        df_f, extra_col="فرق عدد ايام من اخر متابعة",
                        extra_label=day_col_label, extra_agg="sum"
                    )
                    if days_summary.empty:
                        st.markdown('<div class="empty-state">لا توجد بيانات مطابقة لهذا الفلتر</div>', unsafe_allow_html=True)
                    else:
                        st.dataframe(
                            style_summary(days_summary, extra_label=day_col_label),
                            use_container_width=True, hide_index=True
                        )

                with tab_raw:
                    st.markdown("#### 🗂️ أصل البيانات - الإهمال")
                    st.dataframe(df_f, use_container_width=True, hide_index=True)

                # ---------------------------
                # الشارتات
                # ---------------------------
                st.markdown('<div class="section-title">📈 الرسوم البيانية</div>', unsafe_allow_html=True)

                SNB_GREEN = "#00693E"
                SNB_GOLD = "#C9A227"
                SNB_RED = "#A33A3A"

                # النص بيبقى أبيض في الوضع الداكن (Dark) وأسود في الوضع الفاتح (Light)
                def get_theme_text_color():
                    try:
                        base = st.get_option("theme.base")
                    except Exception:
                        base = None
                    return "#FFFFFF" if base == "dark" else "#111827"

                AXIS_TEXT_COLOR = get_theme_text_color()
                GRID_COLOR = "#374151" if AXIS_TEXT_COLOR == "#FFFFFF" else "#f1f5f3"
                LABEL_FONT = dict(size=14, family="Tajawal", color=AXIS_TEXT_COLOR)

                def style_fig(fig, angle=-20):
                    fig.update_xaxes(tickfont=dict(size=13, family="Tajawal", color=AXIS_TEXT_COLOR))
                    fig.update_yaxes(tickfont=dict(size=12, family="Tajawal", color=AXIS_TEXT_COLOR), gridcolor=GRID_COLOR)
                    fig.update_layout(
                        height=400, xaxis_tickangle=angle, margin=dict(t=20, b=10, l=10, r=10),
                        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Tajawal"),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                    )
                    return fig

                chart_col1, chart_col2 = st.columns(2)

                with chart_col1:
                    st.markdown('<div class="chart-card"><div class="chart-card-title">مبلغ المديونية لكل مشرف</div>', unsafe_allow_html=True)
                    sup_amount = (
                        df_f.groupby(supervisor_col, dropna=False)["Net Amount"].sum()
                        .reset_index().rename(columns={supervisor_col: "المشرف", "Net Amount": "مبلغ المديونية"})
                    )
                    if not sup_amount.empty:
                        fig1 = px.bar(
                            sup_amount, x="المشرف", y="مبلغ المديونية", text="مبلغ المديونية",
                            color_discrete_sequence=[SNB_RED], template="plotly_white"
                        )
                        fig1.update_traces(texttemplate="<b>%{text:,.0f}</b>", textposition="outside", textfont=LABEL_FONT)
                        st.plotly_chart(style_fig(fig1), use_container_width=True)
                    else:
                        st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                with chart_col2:
                    st.markdown(f'<div class="chart-card"><div class="chart-card-title">متوسط {day_col_label} لكل مشرف</div>', unsafe_allow_html=True)
                    days_by_sup = (
                        df_f.groupby(supervisor_col, dropna=False)["فرق عدد ايام من اخر متابعة"].mean()
                        .reset_index().rename(columns={
                            supervisor_col: "المشرف",
                            "فرق عدد ايام من اخر متابعة": "متوسط عدد الأيام"
                        })
                    )
                    if not days_by_sup.empty:
                        fig3 = px.bar(
                            days_by_sup, x="المشرف", y="متوسط عدد الأيام", text="متوسط عدد الأيام",
                            color_discrete_sequence=[SNB_GREEN], template="plotly_white"
                        )
                        fig3.update_traces(texttemplate="<b>%{text:.1f}</b>", textposition="outside", textfont=LABEL_FONT)
                        st.plotly_chart(style_fig(fig3), use_container_width=True)
                    else:
                        st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('<div class="chart-card"><div class="chart-card-title">أعلى 15 موظف بعدد الحسابات (الإهمال)</div>', unsafe_allow_html=True)
                emp_count = (
                    df_f.groupby("Salesperson", dropna=False)["Account Number"].nunique()
                    .reset_index(name="عدد الحسابات")
                    .sort_values("عدد الحسابات", ascending=False).head(15)
                )
                if not emp_count.empty:
                    fig2 = px.bar(
                        emp_count, x="Salesperson", y="عدد الحسابات", text="عدد الحسابات",
                        color_discrete_sequence=[SNB_GOLD], template="plotly_white"
                    )
                    fig2.update_traces(texttemplate="<b>%{text}</b>", textposition="outside", textfont=LABEL_FONT)
                    st.plotly_chart(style_fig(fig2, angle=-30), use_container_width=True)
                else:
                    st.markdown('<div class="empty-state">لا توجد بيانات لعرضها</div>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

          except KeyError as e:
            show_error(e)
          except Exception as e:
            show_error(e)
        else:
            st.markdown('<div class="empty-state">⬆️ ارفع ملف المحفظة عشان يبدأ التحليل</div>', unsafe_allow_html=True)

   
    
         

# ======================
# PAGE 4 - التغطية
# ======================
elif page == "التغطية":
    st.subheader("🌐 التغطية")
    st.write("نسب التغطية")


# ======================
# PAGE 5 - النشاط
# ======================
# ============================================================
# ملاحظة مهمة قبل الكود:
# مفيش مكتبة بايثون (openpyxl / xlsxwriter) بتقدر تعمل
# PivotTable حقيقي (اللي بتسحب فيه الحقول بنفسك جوه إكسيل)
# من غير إكسيل نفسه شغال (COM) — ودة مش متاح على Streamlit Cloud
# لأنه Linux وملوش Excel.
# البديل العملي اللي هعمله: شيت "Pivot" منسق كـ Excel Table
# حقيقي (فيه Filter + Sort جاهزين) وبيدي بالظبط الأرقام اللي
# طلبتها (ناجحة / مغطاة لكل محصل). لو عايز Pivot تفاعلي 100%
# تقدر تعمل Insert > PivotTable على الشيت دة من جوه إكسيل نفسه
# في ثانية لأن الداتا هتبقى جاهزة كـ Table.
# ============================================================

# ============================================================
# محتاج تضيف الاستيراد ده فوق مع باقي الـ imports (لو مش موجود):
#   import plotly.express as px
# ومحتاج تضيف plotly في requirements.txt
# ============================================================
# ملاحظة عن الألوان: مفيش ملف ألوان رسمي متاح للعامة من البنك
# الأهلي السعودي (SNB)، فاستخدمت باليتة أخضر غامق/ذهبي قريبة
# من هوية البنوك السعودية بشكل عام (مش شعار أو أصول مسجلة
# ملكية للبنك) — لو عندك الـ Brand Guidelines بتاعتهم ابعتلي
# الأكواد بالظبط وأظبطها.
# ============================================================

elif page == "النشاط":

    st.subheader("⚡ النشاط")

    uploaded_file = st.file_uploader(
        "رفع ملف واحد فقط",
        type=["xlsx", "xls"],
        accept_multiple_files=False
    )

    if uploaded_file:

        # =========================
        # الخطوة 1: هل فيه بريك؟
        # =========================

        if "break_choice" not in st.session_state:
            st.session_state.break_choice = None

        if st.session_state.break_choice is None:

            st.markdown("### هل يوجد بريك؟")
            col1, col2 = st.columns(2)

            with col1:
                if st.button("Yes", use_container_width=True):
                    st.session_state.break_choice = "yes"
                    st.rerun()

            with col2:
                if st.button("No", use_container_width=True):
                    st.session_state.break_choice = "no"
                    st.rerun()

            st.stop()

        if st.button("🔄 تغيير الاختيار (فيه بريك / مفيش)"):
            st.session_state.break_choice = None
            st.session_state.pop("processed_output", None)
            st.session_state.pop("processed_df", None)
            st.rerun()

        has_break = st.session_state.break_choice == "yes"
        break_start_minutes = None
        break_end_minutes = None

        if has_break:
            st.markdown("### 🕐 إعدادات البريك")
            col1, col2 = st.columns(2)

            with col1:
                break_start = st.time_input("البريك يبدأ الساعة", value=time(12, 0))
            with col2:
                break_duration = st.number_input("مدة البريك بالدقائق", min_value=1, value=30, step=5)

            break_start_minutes = break_start.hour * 60 + break_start.minute
            break_end_minutes = break_start_minutes + int(break_duration)

            st.info(
                f"البريك من {break_start.strftime('%H:%M')} إلى "
                f"{(datetime.combine(date.today(), break_start) + timedelta(minutes=break_duration)).strftime('%H:%M')}"
            )
        else:
            st.info("تم اختيار: لا يوجد بريك.")

        start_button = st.button("🚀 ابدأ التصنيف", type="primary")

        # =========================
        # المعالجة (مرة واحدة فقط) + تخزين النتيجة في session_state
        # =========================

        if start_button:

            df = pd.read_excel(uploaded_file)
            df = df.iloc[1:].reset_index(drop=True)

            required_columns = ["Notes", "Created on", "Collector"]
            missing_columns = [c for c in required_columns if c not in df.columns]
            if missing_columns:
                st.error(f"الأعمدة التالية غير موجودة في الملف: {', '.join(missing_columns)}")
                st.stop()

            duplicate_columns = ["Collector", "ID Number", "Notes"]
            missing_duplicate_columns = [c for c in duplicate_columns if c not in df.columns]
            if missing_duplicate_columns:
                st.error(f"الأعمدة التالية غير موجودة: {', '.join(missing_duplicate_columns)}")
                st.stop()

            if "Final State" not in df.columns:
                st.error("لا يوجد عمود Final State في الملف")
                st.stop()

            progress_bar = st.progress(0)
            status = st.empty()

            predictions = []
            probabilities = []
            total = len(df)

            for i, text in enumerate(df["Notes"]):
                label, prob = predict_text(text)
                predictions.append(label)
                probabilities.append(prob)
                progress = (i + 1) / total
                progress_bar.progress(progress)
                status.text(f"{int(progress * 100)}% ({i + 1}/{total})")

            notes_index = df.columns.get_loc("Notes")
            df.insert(notes_index + 1, "التصنيف", predictions)
            df.insert(notes_index + 2, "Probability (%)", probabilities)
            df["التصنيف"] = df["التصنيف"].astype(str)

            df["Created on"] = pd.to_datetime(df["Created on"], errors="coerce")
            df = df.drop_duplicates(subset=duplicate_columns, keep="first").reset_index(drop=True)
            df = df.sort_values(["Collector", "Created on"]).reset_index(drop=True)

            df["فرق التوقيت"] = (
                df.groupby("Collector")["Created on"].diff().dt.total_seconds().div(60)
            )

            def calculate_wasted(row):
                call_time = row["Created on"]
                classification = row["التصنيف"]
                time_diff = row["فرق التوقيت"]

                if pd.isna(time_diff) or pd.isna(call_time):
                    return 0
                if time_diff < 0:
                    return 0

                if has_break:
                    call_minutes = call_time.hour * 60 + call_time.minute
                    if break_start_minutes <= call_minutes < break_end_minutes:
                        return 0

                if classification == "1":
                    wasted = time_diff - 20
                elif classification == "0":
                    wasted = time_diff - 5
                else:
                    wasted = 0

                return max(wasted, 0)

            df["وقت مهدر"] = df.apply(calculate_wasted, axis=1)

            cols = list(df.columns)
            prob_index = cols.index("Probability (%)")
            for col in ["فرق التوقيت", "وقت مهدر"]:
                cols.remove(col)
            prob_index = cols.index("Probability (%)")
            cols.insert(prob_index + 1, "فرق التوقيت")
            cols.insert(prob_index + 2, "وقت مهدر")
            df = df[cols]

            collector_summary = (
                df.groupby("Collector", as_index=False)["وقت مهدر"]
                .sum().sort_values("وقت مهدر", ascending=False).reset_index(drop=True)
            )
            calls_over_30 = df[df["فرق التوقيت"] > 30].sort_values("فرق التوقيت", ascending=False)
            calls_under_1 = df[df["فرق التوقيت"] < 1].sort_values("فرق التوقيت", ascending=True)

            first_activity = (
                df.dropna(subset=["Created on"]).sort_values(["Collector", "Created on"])
                .groupby("Collector", as_index=False).first()[["Collector", "Created on", "التصنيف", "Probability (%)"]]
                .rename(columns={"Created on": "وقت أول إفادة"})
            )
            last_activity = (
                df.dropna(subset=["Created on"]).sort_values(["Collector", "Created on"])
                .groupby("Collector", as_index=False).last()[["Collector", "Created on", "التصنيف", "Probability (%)", "Final State", "Notes"]]
                .rename(columns={"Created on": "وقت آخر إفادة"})
            )

            final_state_summary = df["Final State"].fillna("Blank").value_counts().reset_index()
            final_state_summary.columns = ["Final State", "عدد الحالات"]

            pivot_summary = (
                df.groupby(["Collector", "التصنيف"]).size().unstack(fill_value=0)
                .rename(columns={"1": "مكالمات ناجحة", "0": "مكالمات غير ناجحة"})
            )
            for needed_col in ["مكالمات ناجحة", "مكالمات غير ناجحة"]:
                if needed_col not in pivot_summary.columns:
                    pivot_summary[needed_col] = 0
            pivot_summary = pivot_summary[["مكالمات ناجحة", "مكالمات غير ناجحة"]]
            pivot_summary["إجمالي المكالمات المغطاة"] = pivot_summary.sum(axis=1)
            pivot_summary["نسبة النجاح %"] = (
                (pivot_summary["مكالمات ناجحة"] / pivot_summary["إجمالي المكالمات المغطاة"] * 100).round(1)
            )
            pivot_summary = pivot_summary.reset_index().sort_values("إجمالي المكالمات المغطاة", ascending=False).reset_index(drop=True)

            first_after_break = None
            if has_break:
                after_break = df[df["Created on"].notna()].copy()
                after_break["وقت بالدقائق"] = after_break["Created on"].dt.hour * 60 + after_break["Created on"].dt.minute
                after_break = after_break[after_break["وقت بالدقائق"] >= break_end_minutes]
                first_after_break = (
                    after_break.sort_values(["Collector", "Created on"]).groupby("Collector", as_index=False).first()
                    [["Collector", "Created on", "التصنيف", "Probability (%)"]].rename(columns={"Created on": "وقت أول إفادة بعد البريك"})
                )
                first_after_break["حالة النجاح"] = first_after_break["التصنيف"].apply(
                    lambda v: "ناجحة" if v == "1" else ("غير ناجحة" if v == "0" else v)
                )

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="النشاط")
                pivot_summary.to_excel(writer, index=False, sheet_name="Pivot - أداء المحصلين")
                collector_summary.to_excel(writer, index=False, sheet_name="إجمالي الوقت المهدر")
                calls_over_30.to_excel(writer, index=False, sheet_name="مكالمات +30 دقيقة")
                calls_under_1.to_excel(writer, index=False, sheet_name="مكالمات أقل من دقيقة")
                first_activity.to_excel(writer, index=False, sheet_name="أول إفادة")
                last_activity.to_excel(writer, index=False, sheet_name="آخر إفادة")
                if first_after_break is not None:
                    first_after_break.to_excel(writer, index=False, sheet_name="أول إفادة بعد البريك")
                final_state_summary.to_excel(writer, index=False, sheet_name="Final State")

                workbook = writer.book
                header_fill = PatternFill(fill_type="solid", fgColor="073259")
                header_font = Font(color="FFFFFF", bold=True)
                thin_side = Side(style="thin", color="000000")
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                for ws in workbook.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    for column_cells in ws.columns:
                        max_length = 0
                        column_letter = column_cells[0].column_letter
                        for cell in column_cells:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
                    ws.row_dimensions[1].height = 25

                pivot_ws = workbook["Pivot - أداء المحصلين"]
                last_col_letter = pivot_ws.cell(row=1, column=pivot_ws.max_column).column_letter
                table_range = f"A1:{last_col_letter}{pivot_ws.max_row}"
                excel_table = Table(displayName="PivotAdaaAlMohaseleen", ref=table_range)
                excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                pivot_ws.add_table(excel_table)

            output.seek(0)

            st.session_state.processed_output = output.getvalue()
            st.session_state.processed_df = df.copy()

            status.empty()
            progress_bar.empty()
            st.success("تم الانتهاء")

        if "processed_output" in st.session_state:
            st.download_button(
                "تحميل الملف",
                st.session_state.processed_output,
                file_name="activity.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ============================================================
        # الداشبورد (4 Cards + 4 Charts) — بيظهر بعد التصنيف مباشرة
        # ============================================================

# ============================================================
# ده بديل قسم "الداشبورد" بالكامل (من سطر SNB_GREEN = ...
# لحد آخر سطر في الملف اللي فاتت). امسح القسم القديم وحط
# ده مكانه، والباقي (البريك + التصنيف + تحميل الإكسيل) زي ما هو.
# ============================================================

        if "processed_df" in st.session_state:

            data = st.session_state.processed_df

            SNB_GREEN = "#00693E"
            SNB_GREEN_DARK = "#024930"
            SNB_GOLD = "#C9A227"
            SNB_RED = "#A33A3A"

            st.markdown("---")
            st.markdown("## 📊 داشبورد النشاط")

            # ---------------------------
            # السلايسرات (محصل / مشرف)
            # ---------------------------

            possible_supervisor_cols = ["Sales Team", "Supervisor", "المشرف", "Team Leader", "TL", "Manager", "Supervisor Name"]
            supervisor_col = next((c for c in possible_supervisor_cols if c in data.columns), None)

            filter_col1, filter_col2 = st.columns(2)

            with filter_col1:
                selected_collectors = st.multiselect(
                    "فلترة حسب المحصل",
                    options=sorted(data["Collector"].dropna().unique().tolist()),
                    default=[]
                )

            with filter_col2:
                if supervisor_col:
                    selected_supervisors = st.multiselect(
                        "فلترة حسب المشرف",
                        options=sorted(data[supervisor_col].dropna().unique().tolist()),
                        default=[]
                    )
                else:
                    selected_supervisors = []
                    st.caption(
                        "⚠️ مفيش عمود واضح للمشرف في الملف المرفوع "
                        "(بحثت عن: Supervisor / المشرف / Team Leader). "
                        "لو الملف فيه عمود بالاسم دة، ابعتلي اسمه بالظبط عشان أظبط الفلتر."
                    )

            filtered = data.copy()
            if selected_collectors:
                filtered = filtered[filtered["Collector"].isin(selected_collectors)]
            if supervisor_col and selected_supervisors:
                filtered = filtered[filtered[supervisor_col].isin(selected_supervisors)]

            if filtered.empty:
                st.warning("مفيش بيانات مطابقة للفلتر المختار.")
                st.stop()

            # ---------------------------
            # حساب مقاييس كل محصل (مستخدمة في الـ Cards والـ Charts)
            # ---------------------------

            collector_agg = filtered.groupby("Collector").agg(
                المكالمات_المغطاة=("Collector", "size"),
                المكالمات_الناجحة=("التصنيف", lambda s: (s == "1").sum()),
                الوقت_المهدر=("وقت مهدر", "sum"),
            ).reset_index()

            def normalize_0_100(s):
                if s.max() == s.min():
                    return pd.Series([100.0] * len(s), index=s.index)
                return (s - s.min()) / (s.max() - s.min()) * 100

            collector_agg["_covered_norm"] = normalize_0_100(collector_agg["المكالمات_المغطاة"])
            collector_agg["_success_norm"] = normalize_0_100(collector_agg["المكالمات_الناجحة"])
            collector_agg["_wasted_norm"] = normalize_0_100(collector_agg["الوقت_المهدر"])

            # سكور الفاعلية: 35% تغطية + 35% نجاح + 30% (100 - وقت مهدر)
            # يعني اللي بيغطي مكالمات كتير وناجح فيها وبيضيع وقت أقل بيطلع فوق
            collector_agg["سكور الفاعلية"] = (
                collector_agg["_covered_norm"] * 0.35
                + collector_agg["_success_norm"] * 0.35
                + (100 - collector_agg["_wasted_norm"]) * 0.30
            ).round(1)

            collector_agg = collector_agg.sort_values("سكور الفاعلية", ascending=False).reset_index(drop=True)
            top_performer = collector_agg.iloc[0]

            # ---------------------------
            # الـ 4 Cards
            # ---------------------------

            total_covered = len(filtered)
            total_successful = int((filtered["التصنيف"] == "1").sum())
            
            # Average وقت مهدر
            avg_wasted = filtered["وقت مهدر"].mean()
            avg_wasted = 0 if pd.isna(avg_wasted) else avg_wasted
            
            
            def render_card(col, title, value, subtitle=""):
                with col:
                    st.markdown(
                        f'<div style="background: linear-gradient(135deg, {SNB_GREEN} 0%, {SNB_GREEN_DARK} 100%); '
                        f'border-right: 5px solid {SNB_GOLD}; border-radius: 14px; padding: 18px; color: white; '
                        f'text-align: center; box-shadow: 0 4px 14px rgba(0,0,0,0.15); min-height: 130px;">'
                        f'<div style="font-size:13px; opacity:0.85; margin-bottom:8px;">{title}</div>'
                        f'<div style="font-size:24px; font-weight:800; overflow-wrap:anywhere;">{value}</div>'
                        f'<div style="font-size:11px; opacity:0.75; margin-top:6px;">{subtitle}</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
            
            
            c1, c2, c3, c4 = st.columns(4)
            
            render_card(
                c1,
                "عدد المكالمات المغطاة",
                f"{total_covered:,}"
            )
            
            render_card(
                c2,
                "عدد المكالمات الناجحة",
                f"{total_successful:,}",
                f"{(total_successful / total_covered * 100 if total_covered else 0):.1f}% من المغطاة"
            )
            
            render_card(
                c3,
                "متوسط الوقت المهدر (AVG)",
                f"{avg_wasted:.1f} د"
            )
            
            render_card(
                c4,
                "أكتر محصل فاعلية",
                f"{top_performer['Collector']}",
                f"سكور {top_performer['سكور الفاعلية']:.0f}/100"
            )
            
            st.caption(
                "سكور الفاعلية بيوازن بين: عدد المكالمات المغطاة، عدد المكالمات الناجحة، "
                "ووقت الهدر (كل ما الوقت المهدر أقل كل ما السكور أعلى) — كل عامل بوزن نسبي "
                "مقارنة بباقي المحصلين في نفس الفلتر الحالي."
            )

            st.markdown("")

            with st.expander("📋 آخر نشاط لكل محصل"):
                last_per_collector = (
                    filtered.dropna(subset=["Created on"])
                    .sort_values(["Collector", "Created on"])
                    .groupby("Collector", as_index=False)
                    .last()
                )
                show_cols = [c for c in ["Collector", "Created on", "التصنيف", "Probability (%)", "Final State", "Notes"] if c in last_per_collector.columns]
                st.dataframe(last_per_collector[show_cols], use_container_width=True, hide_index=True)

            with st.expander("🏆 ترتيب المحصلين حسب سكور الفاعلية"):
                st.dataframe(
                    collector_agg[["Collector", "المكالمات_المغطاة", "المكالمات_الناجحة", "الوقت_المهدر", "سكور الفاعلية"]],
                    use_container_width=True, hide_index=True
                )

            st.markdown("")

            # ---------------------------
            # الـ 4 Charts — كل شارت ياخد مساحته كاملة (مش نص الشاشة)
            # ---------------------------

            base_font = dict(size=14)

            # 1) مكالمات ناجحة/غير ناجحة لكل محصل — Bar أفقي عشان الأسماء تبقى واضحة
            st.markdown("#### مكالمات ناجحة مقابل غير ناجحة لكل محصل")

            chart1_data = filtered.groupby(["Collector", "التصنيف"]).size().reset_index(name="عدد المكالمات")
            chart1_data["الحالة"] = chart1_data["التصنيف"].map({"1": "ناجحة", "0": "غير ناجحة"}).fillna(chart1_data["التصنيف"])

            n_collectors = filtered["Collector"].nunique()
            fig1_height = max(400, 55 * n_collectors)

            fig1 = px.bar(
                chart1_data, y="Collector", x="عدد المكالمات", color="الحالة",
                orientation="h", barmode="stack", text="عدد المكالمات",
                color_discrete_map={"ناجحة": SNB_GREEN, "غير ناجحة": SNB_GOLD}
            )
            fig1.update_traces(textposition="inside", textfont=dict(color="white", size=13))
            fig1.update_layout(
                height=fig1_height, font=base_font, legend_title_text="",
                margin=dict(t=10, b=10),
                yaxis={"categoryorder": "total ascending"}
            )
            st.plotly_chart(fig1, use_container_width=True)

            # 2) توزيع Final State — Column chart بدل الـ Pie
            st.markdown("#### توزيع Final State")

            fs_data = filtered["Final State"].fillna("Blank").value_counts().reset_index()
            fs_data.columns = ["Final State", "عدد"]
            fs_data = fs_data.sort_values("عدد", ascending=False)

            if len(fs_data) > 12:
                top_fs = fs_data.iloc[:11]
                other_sum = fs_data.iloc[11:]["عدد"].sum()
                fs_data = pd.concat([top_fs, pd.DataFrame({"Final State": ["أخرى"], "عدد": [other_sum]})], ignore_index=True)

            fig2 = px.bar(
                fs_data, x="Final State", y="عدد", text="عدد",
                color_discrete_sequence=[SNB_GREEN]
            )
            fig2.update_traces(textposition="outside", textfont=dict(size=13))
            fig2.update_layout(
                height=450, font=base_font, margin=dict(t=20, b=10),
                xaxis_tickangle=-30
            )
            st.plotly_chart(fig2, use_container_width=True)

            # 3) الوقت المهدر لكل محصل — Bar
            st.markdown("#### الوقت المهدر لكل محصل")

            wasted_data = (
                filtered.groupby("Collector", as_index=False)["وقت مهدر"]
                .sum().sort_values("وقت مهدر", ascending=False)
            )
            fig3_height = max(400, 45 * len(wasted_data))

            fig3 = px.bar(
                wasted_data, x="وقت مهدر", y="Collector", orientation="h",
                text="وقت مهدر", color_discrete_sequence=[SNB_RED]
            )
            fig3.update_traces(texttemplate="%{text:.0f}", textposition="outside", textfont=dict(size=13))
            fig3.update_layout(
                height=fig3_height, font=base_font, margin=dict(t=10, b=10),
                yaxis={"categoryorder": "total ascending"}
            )
            st.plotly_chart(fig3, use_container_width=True)

            # 4) توزيع المكالمات عبر ساعات اليوم
            st.markdown("#### توزيع المكالمات عبر ساعات اليوم")

            hour_data = filtered.dropna(subset=["Created on"]).copy()
            hour_data["الساعة"] = hour_data["Created on"].dt.hour
            hour_counts = hour_data.groupby("الساعة").size().reset_index(name="عدد المكالمات")

            fig4 = px.bar(
                hour_counts, x="الساعة", y="عدد المكالمات", text="عدد المكالمات",
                color_discrete_sequence=[SNB_GREEN_DARK]
            )
            fig4.update_traces(textposition="outside", textfont=dict(size=13))
            fig4.update_layout(
                height=420, font=base_font, margin=dict(t=20, b=10),
                xaxis=dict(dtick=1)
            )
            st.plotly_chart(fig4, use_container_width=True)


            # ============================================================
# 1) تعديل بسيط في تعريف عمود المشرف — ضيف "Sales Team" أول
#    القايمة عشان يتكشف هو الأول قبل أي اسم تاني:
#
#    possible_supervisor_cols = ["Sales Team", "Supervisor", "المشرف",
#                                 "Team Leader", "TL", "Manager", "Supervisor Name"]
#
#    (بدّل السطر القديم بده في نفس مكانه جوه قسم الداشبورد)
# ============================================================
# 2) القسم ده يتحط في الآخر، بعد آخر chart (توزيع المكالمات عبر
#    ساعات اليوم) وقبل ما يخلص الـ if "processed_df" in st.session_state:
# ============================================================

            # ---------------------------
            # تحليل العملاء (Debtor / Customer Account Number)
            # ---------------------------

# ============================================================
# 1) تعديل بسيط في تعريف عمود المشرف — ضيف "Sales Team" أول
#    القايمة عشان يتكشف هو الأول قبل أي اسم تاني:
#
#    possible_supervisor_cols = ["Sales Team", "Supervisor", "المشرف",
#                                 "Team Leader", "TL", "Manager", "Supervisor Name"]
#
#    (بدّل السطر القديم بده في نفس مكانه جوه قسم الداشبورد)
# ============================================================
# 2) القسم ده يتحط في الآخر، بعد آخر chart (توزيع المكالمات عبر
#    ساعات اليوم) وقبل ما يخلص الـ if "processed_df" in st.session_state:
# ============================================================

            # ---------------------------
            # تحليل العملاء (Debtor / Customer Account Number)
            # ---------------------------

            st.markdown("---")
            st.markdown("## 👥 تحليل العملاء")

            debtor_col = "Debtor"
            account_col = "Customer Account Number"
            missing_client_cols = [c for c in [debtor_col, account_col] if c not in filtered.columns]

            if missing_client_cols:
                st.warning(
                    f"الأعمدة دي مش موجودة في الملف المرفوع: {', '.join(missing_client_cols)} — "
                    "قسم تحليل العملاء مش هيظهر. تأكد من اسم العمود بالظبط في الشيت لو الاسم مختلف."
                )
            else:

                # كل موظف تابع كام عميل (Debtor) وكام رقم حساب (Customer Account Number)
                client_coverage = (
                    filtered.groupby("Collector")
                    .agg(
                        عدد_العملاء=(debtor_col, "nunique"),
                        عدد_أرقام_الحسابات=(account_col, "nunique"),
                        عدد_الإفادات=(debtor_col, "count"),
                    )
                    .reset_index()
                    .sort_values("عدد_العملاء", ascending=False)
                )

                st.markdown("#### كل موظف تابع كام عميل وكام رقم حساب")
                st.dataframe(client_coverage, use_container_width=True, hide_index=True)

                # الموظف اللي كرر نفس العميل أكتر من 5 مرات (أكتر من 5 إفادات على نفس الـ Debtor)
                repeat_contacts = (
                    filtered.groupby(["Collector", debtor_col])
                    .size()
                    .reset_index(name="عدد الإفادات على نفس العميل")
                )
                repeat_contacts = repeat_contacts[repeat_contacts["عدد الإفادات على نفس العميل"] >= 2]
                repeat_contacts = repeat_contacts.sort_values(
                    "عدد الإفادات على نفس العميل", ascending=False
                ).reset_index(drop=True)

                st.markdown("#### موظفين كلموا نفس العميل أكتر من مرة")

                if repeat_contacts.empty:
                    st.info("مفيش أي موظف كلم نفس العميل أكتر من مرة في البيانات الحالية.")
                else:
                    st.dataframe(repeat_contacts, use_container_width=True, hide_index=True)

                # =========================
                # زرار تحميل منفصل لتحليل العملاء
                # =========================

                client_output = BytesIO()
                with pd.ExcelWriter(client_output, engine="openpyxl") as writer:
                    client_coverage.to_excel(writer, index=False, sheet_name="تغطية العملاء لكل موظف")
                    repeat_contacts.to_excel(writer, index=False, sheet_name="تكرار نفس العميل")

                    workbook = writer.book
                    header_fill = PatternFill(fill_type="solid", fgColor="073259")
                    header_font = Font(color="FFFFFF", bold=True)
                    thin_side = Side(style="thin", color="000000")
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                    for ws in workbook.worksheets:
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        for column_cells in ws.columns:
                            max_length = 0
                            column_letter = column_cells[0].column_letter
                            for cell in column_cells:
                                if cell.value is not None:
                                    max_length = max(max_length, len(str(cell.value)))
                            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
                        ws.row_dimensions[1].height = 25

                client_output.seek(0)

                st.download_button(
                    "⬇️ تحميل تحليل العملاء",
                    client_output.getvalue(),
                    file_name="client_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ======================
# PAGE 6 - باقي الصفحات (مختصر)
# ======================
elif page == "التوزيع":

    st.subheader("📦 توزيع المحافظ")
    uploaded_file = st.file_uploader(
        "ارفع ملف Excel",
        type=["xlsx"]
    )
    
   
        





elif page == "اخطاء الحالات":
    st.subheader("❌ اخطاء الحالات")

elif page == "الاوتودايلر":
    st.subheader("📞 الاوتودايلر")

elif page == "التدوير":

    import pandas as pd
    import numpy as np
    import plotly.express as px
    import traceback
    from io import BytesIO

    # ============================================================
    # 🎨 نظام تصميم مودرن شامل (نفس عائلة التصميم، بلمسة لون خاصة بالتدوير)
    # ============================================================
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800&display=swap');

        html, body, [class*="css"] {
            font-family: 'Tajawal', sans-serif;
        }

        .main .block-container {
            padding-top: 1.2rem;
            padding-bottom: 3rem;
        }

        /* ===== الهيدر الرئيسي ===== */
        .rotation-header {
            position: relative;
            overflow: hidden;
            background: linear-gradient(120deg, #0d2d4a 0%, #155a8a 50%, #1c76b3 100%);
            padding: 32px 34px;
            border-radius: 20px;
            margin-bottom: 26px;
            box-shadow: 0 10px 30px rgba(21,90,138,0.28);
        }
        .rotation-header::after {
            content: "";
            position: absolute;
            top: -60px; left: -40px;
            width: 220px; height: 220px;
            background: radial-gradient(circle, rgba(255,255,255,0.10) 0%, rgba(255,255,255,0) 70%);
            border-radius: 50%;
        }
        .rotation-header h1 {
            color: #fff;
            margin: 0;
            font-size: 27px;
            font-weight: 800;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .rotation-header p {
            color: #d3e9f7;
            margin: 8px 0 0 0;
            font-size: 14.5px;
            font-weight: 500;
        }
        .header-badge {
            display: inline-block;
            background: rgba(255,255,255,0.14);
            color: #fff;
            font-size: 12px;
            font-weight: 700;
            padding: 4px 12px;
            border-radius: 999px;
            margin-top: 12px;
            border: 1px solid rgba(255,255,255,0.25);
        }

        /* ===== بطاقات KPI ===== */
        .kpi-grid-3 {
            display: grid;
            grid-template-columns: 1fr 1fr 1fr;
            gap: 16px;
            margin-bottom: 6px;
        }
        .kpi-card {
            position: relative;
            background: #ffffff;
            border-radius: 18px;
            padding: 20px 22px;
            box-shadow: 0 6px 20px rgba(17,24,39,0.07);
            border: 1px solid #eef1ef;
            border-left: 6px solid #155a8a;
            overflow: hidden;
        }
        .kpi-card.ok { border-left-color: #00693E; }
        .kpi-card::before {
            content: "";
            position: absolute;
            top: 0; right: 0;
            width: 90px; height: 90px;
            background: radial-gradient(circle, rgba(21,90,138,0.10) 0%, rgba(21,90,138,0) 70%);
        }
        .kpi-icon {
            width: 40px; height: 40px;
            border-radius: 12px;
            display: flex; align-items: center; justify-content: center;
            font-size: 19px;
            background: #e5f0f8;
        }
        .kpi-card.ok .kpi-icon { background: #e7f4ec; }
        .kpi-label {
            font-size: 13px;
            color: #6b7280;
            font-weight: 700;
            margin-top: 12px;
        }
        .kpi-value {
            font-size: 30px;
            font-weight: 800;
            color: #0f172a;
            margin-top: 2px;
        }
        .kpi-sub {
            font-size: 12px;
            color: #9aa4b2;
            font-weight: 600;
            margin-top: 4px;
        }

        /* ===== عناوين الأقسام ===== */
        .section-title {
            display: flex;
            align-items: center;
            gap: 10px;
            background: linear-gradient(90deg, #f1f7fb 0%, #ffffff 100%);
            border-radius: 12px;
            padding: 12px 18px;
            margin: 26px 0 14px 0;
            border-right: 5px solid #155a8a;
            font-weight: 800;
            font-size: 16.5px;
            color: #0d2d4a;
        }

        /* ===== بطاقة تحيط بالشارت ===== */
        .chart-card {
            background: #ffffff;
            border-radius: 18px;
            padding: 14px 16px 4px 16px;
            border: 1px solid #eef1ef;
            box-shadow: 0 4px 14px rgba(17,24,39,0.05);
            margin-bottom: 18px;
        }
        .chart-card-title {
            font-weight: 800;
            font-size: 14.5px;
            color: #0f172a;
            margin-bottom: 4px;
        }

        /* ===== رفع الملف ===== */
        div[data-testid="stFileUploader"] {
            border: 2px dashed #155a8a44;
            border-radius: 16px;
            padding: 8px;
            background: #f8fbfd;
        }

        /* ===== أزرار التحميل ===== */
        div[data-testid="stDownloadButton"] button {
            border-radius: 12px !important;
            font-weight: 700 !important;
            border: 1px solid #d8e6f0 !important;
            transition: all 0.15s ease-in-out;
        }
        div[data-testid="stDownloadButton"] button:hover {
            border-color: #155a8a !important;
            color: #155a8a !important;
            transform: translateY(-1px);
        }

        /* ===== حالة فارغة ===== */
        .empty-state {
            text-align: center;
            padding: 26px 10px;
            color: #9aa4b2;
            font-weight: 600;
            font-size: 14px;
        }

        /* ===== صندوق الأخطاء ===== */
        .error-box {
            background: #fdecea;
            border: 1px solid #f5c2c0;
            border-radius: 14px;
            padding: 18px 20px;
            color: #7a1f1a;
            font-weight: 600;
            line-height: 2;
        }
        .error-box code {
            background: #fbe0de;
            padding: 2px 7px;
            border-radius: 6px;
            font-weight: 700;
        }
        .error-title {
            font-size: 16px;
            font-weight: 800;
            margin-bottom: 6px;
        }

        /* ===== صندوق تأكيد النجاح ===== */
        .success-box {
            background: #eaf6ef;
            border: 1px solid #c9e9d5;
            border-radius: 14px;
            padding: 14px 18px;
            color: #0f3d2e;
            font-weight: 700;
            margin-bottom: 14px;
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="rotation-header">
        <h1>🔄 التدوير</h1>
        <p>إعادة توزيع العملاء على المحصلين بحيث لا يحتفظ أي عميل بمحصله القديم، مع الحفاظ على نفس عدد العملاء ومتبقي المديونية لكل محصل قدر الإمكان</p>
        <span class="header-badge">توزيع آلي متوازن + تحسين محلي</span>
    </div>
    """, unsafe_allow_html=True)

    # ============================================================
    # 🛠️ أدوات مساعدة للتحقق من الأعمدة وعرض الأخطاء بدقة
    # ============================================================
    def require_columns(df, required_cols, step_name):
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise KeyError("STEP::" + step_name + "::MISSING::" + "|".join(missing))

    def show_error(exc):
        msg = str(exc)
        if msg.startswith("STEP::"):
            try:
                _, step_name, _, cols_part = msg.split("::", 3)
                cols_list = cols_part.split("|")
                cols_html = "".join([f"<li><code>{c}</code></li>" for c in cols_list])
                st.markdown(f"""
                <div class="error-box">
                    <div class="error-title">❌ حصل خطأ أثناء تنفيذ خطوة: <code>{step_name}</code></div>
                    التفاصيل:
                    <ul>{cols_html}</ul>
                </div>
                """, unsafe_allow_html=True)
            except Exception:
                st.error(f"❌ خطأ غير متوقع: {msg}")
        else:
            st.markdown(f"""
            <div class="error-box">
                <div class="error-title">❌ حصل خطأ غير متوقع</div>
                <b>نوع الخطأ:</b> <code>{type(exc).__name__}</code><br>
                <b>تفاصيل:</b> {msg}
            </div>
            """, unsafe_allow_html=True)
            with st.expander("🔍 تفاصيل تقنية (Traceback)"):
                st.code(traceback.format_exc())

    # ============================================================
    # 🧠 مرحلة التحسين المحلي (Local Search) بعد التوزيع الأولي
    # ============================================================
    def _sample_by_debt(items, cap):
        """
        لو عدد العملاء عند محصل معين كبير جدًا، بناخد عينة موزعة على مستويات
        المديونية المختلفة (مش بس الأعلى) بدل ما نفحص كل الاحتمالات (أداء أسرع).
        """
        if len(items) <= cap:
            return items
        items_sorted = sorted(items, key=lambda x: x["debt"], reverse=True)
        step = len(items_sorted) / cap
        return [items_sorted[int(i * step)] for i in range(cap)]

    def refine_assignment(groups, assignment, current_count, current_debt,
                           target_count, target_debt, collectors,
                           max_passes=30, pool_cap=120):
        """
        تحسين محلي (Local Search) فوق نتيجة التوزيع الأولي (Greedy) لتقليل
        الفرق الكلي بين (قديم/جديد) في عدد العملاء ومتبقي المديونية معًا:

        - Move  : نقل عميل واحد (كل حساباته) لمحصل تاني لو ده بيقلل الانحراف
                  الكلي (بيأثر على العدد والمديونية عند المحصلين المعنيين).
        - Swap  : تبديل عميلين بين محصلين مختلفين. عدد العملاء عند الاتنين
                  بيفضل زي ما هو، وبس بيتحسن توازن متبقي المديونية بينهم.

        بتكرر الاتنين على شكل "passes" لحد ما محدش يقدر يحسن أكتر أو
        نوصل للحد الأقصى لعدد المحاولات.
        """

        def pen(count_val, target_c, debt_val, target_d):
            tc = target_c or 1
            td = target_d or 1.0
            pc = (count_val - target_c) / tc
            pdv = (debt_val - target_d) / td
            return pc * pc + pdv * pdv

        groups_by_collector = {c: [] for c in collectors}
        for g in groups:
            cid = assignment.get(g["id"])
            if cid is not None:
                groups_by_collector[cid].append(g)

        passes_done = 0
        improved = True

        while improved and passes_done < max_passes:
            improved = False
            passes_done += 1

            # ---------- 1) Move: نقل عميل واحد لمحصل أنسب ----------
            for g in groups:
                gid = g["id"]
                c_old = assignment[gid]

                old_pen_old = pen(
                    current_count[c_old], target_count.get(c_old, 0),
                    current_debt[c_old], target_debt.get(c_old, 0.0)
                )

                best_c, best_delta = None, -1e-9

                for c_new in collectors:
                    if c_new == c_old or c_new in g["forbidden"]:
                        continue

                    old_pen_new = pen(
                        current_count[c_new], target_count.get(c_new, 0),
                        current_debt[c_new], target_debt.get(c_new, 0.0)
                    )
                    new_pen_old = pen(
                        current_count[c_old] - 1, target_count.get(c_old, 0),
                        current_debt[c_old] - g["debt"], target_debt.get(c_old, 0.0)
                    )
                    new_pen_new = pen(
                        current_count[c_new] + 1, target_count.get(c_new, 0),
                        current_debt[c_new] + g["debt"], target_debt.get(c_new, 0.0)
                    )

                    delta = (new_pen_old + new_pen_new) - (old_pen_old + old_pen_new)
                    if delta < best_delta:
                        best_delta = delta
                        best_c = c_new

                if best_c is not None:
                    current_count[c_old] -= 1
                    current_debt[c_old] -= g["debt"]
                    current_count[best_c] += 1
                    current_debt[best_c] += g["debt"]
                    assignment[gid] = best_c
                    groups_by_collector[c_old].remove(g)
                    groups_by_collector[best_c].append(g)
                    improved = True

            # ---------- 2) Swap: تبديل عميلين بين محصلين لتوازن المديونية ----------
            for i, c1 in enumerate(collectors):
                for c2 in collectors[i + 1:]:
                    list1 = [g for g in groups_by_collector[c1] if c2 not in g["forbidden"]]
                    list2 = [g for g in groups_by_collector[c2] if c1 not in g["forbidden"]]
                    if not list1 or not list2:
                        continue

                    list1 = _sample_by_debt(list1, pool_cap)
                    list2 = _sample_by_debt(list2, pool_cap)

                    old_pen_pair = pen(
                        current_count[c1], target_count.get(c1, 0),
                        current_debt[c1], target_debt.get(c1, 0.0)
                    ) + pen(
                        current_count[c2], target_count.get(c2, 0),
                        current_debt[c2], target_debt.get(c2, 0.0)
                    )

                    best_pair, best_delta = None, -1e-9

                    for g1 in list1:
                        for g2 in list2:
                            new_debt1 = current_debt[c1] - g1["debt"] + g2["debt"]
                            new_debt2 = current_debt[c2] - g2["debt"] + g1["debt"]
                            new_pen_pair = pen(
                                current_count[c1], target_count.get(c1, 0),
                                new_debt1, target_debt.get(c1, 0.0)
                            ) + pen(
                                current_count[c2], target_count.get(c2, 0),
                                new_debt2, target_debt.get(c2, 0.0)
                            )
                            delta = new_pen_pair - old_pen_pair
                            if delta < best_delta:
                                best_delta = delta
                                best_pair = (g1, g2)

                    if best_pair:
                        g1, g2 = best_pair
                        assignment[g1["id"]] = c2
                        assignment[g2["id"]] = c1
                        current_debt[c1] += (g2["debt"] - g1["debt"])
                        current_debt[c2] += (g1["debt"] - g2["debt"])
                        groups_by_collector[c1].remove(g1)
                        groups_by_collector[c1].append(g2)
                        groups_by_collector[c2].remove(g2)
                        groups_by_collector[c2].append(g1)
                        improved = True

        return passes_done

    st.markdown("""
    الملف المطلوب لازم يحتوي على 4 أعمدة:
    **رقم الهوية** — **اسم المحصل القديم** — **متبقي المديونية** — **رقم الحساب**
    """)

    rotation_file = st.file_uploader(
        "رفع ملف المحفظة",
        type=["xlsx", "xls"],
        key="rotation_file_uploader"
    )

    REQUIRED_ROTATION_COLS = [
        "رقم الهوية", "اسم المحصل القديم", "متبقي المديونية", "رقم الحساب"
    ]

    # ============================================================
    # الدالة الأساسية لإعادة التوزيع - Cached
    # ============================================================
    @st.cache_data(show_spinner="جاري إعادة توزيع المحفظة على المحصلين...")
    def rotate_portfolio(file_bytes):
        try:
            df = pd.read_excel(BytesIO(file_bytes))
        except Exception as e:
            raise KeyError(f"STEP::قراءة ملف الإكسيل::MISSING::{e}")

        require_columns(df, REQUIRED_ROTATION_COLS, "التحقق من الأعمدة المطلوبة")

        df = df.copy()
        df["رقم الهوية"] = df["رقم الهوية"].astype(str).str.strip()
        df["اسم المحصل القديم"] = df["اسم المحصل القديم"].astype(str).str.strip()
        df["رقم الحساب"] = df["رقم الحساب"].astype(str).str.strip()

        df["متبقي المديونية"] = (
            df["متبقي المديونية"].astype(str).str.replace(",", "", regex=False).str.strip()
        )
        df["متبقي المديونية"] = pd.to_numeric(df["متبقي المديونية"], errors="coerce").fillna(0.0)

        collectors = sorted([c for c in df["اسم المحصل القديم"].unique().tolist() if c and c.lower() != "nan"])
        if len(collectors) < 2:
            raise KeyError(
                "STEP::التحقق من عدد المحصلين::MISSING::"
                "لازم يكون في الملف محصلين اثنين على الأقل عشان تدوير المحفظة"
            )

        # ------------------------------------------------------
        # الأهداف الأصلية لكل محصل:
        # عدد العملاء (IDs مميزة) + إجمالي متبقي المديونية (على مستوى الحسابات)
        # ------------------------------------------------------
        target_count = df.groupby("اسم المحصل القديم")["رقم الهوية"].nunique().to_dict()
        target_debt = df.groupby("اسم المحصل القديم")["متبقي المديونية"].sum().to_dict()

        # ------------------------------------------------------
        # تجميع الصفوف حسب رقم الهوية — كل عميل ينتقل ككتلة واحدة
        # ------------------------------------------------------
        groups = []
        for id_val, g in df.groupby("رقم الهوية", sort=False):
            groups.append({
                "id": id_val,
                "debt": float(g["متبقي المديونية"].sum()),
                "forbidden": set(g["اسم المحصل القديم"].unique().tolist()),
            })

        # المجموعات الأكبر (من حيث المديونية) الأول، لتوزيع أفضل
        groups.sort(key=lambda x: x["debt"], reverse=True)

        current_count = {c: 0 for c in collectors}
        current_debt = {c: 0.0 for c in collectors}
        assignment = {}
        unassignable = []

        for grp in groups:
            candidates = [c for c in collectors if c not in grp["forbidden"]]
            if not candidates:
                unassignable.append(str(grp["id"]))
                continue

            def score(c):
                tc = target_count.get(c, 0) or 1
                td = target_debt.get(c, 0.0) or 1.0
                deficit_count = (target_count.get(c, 0) - current_count[c]) / tc
                deficit_debt = (target_debt.get(c, 0.0) - current_debt[c]) / td
                return deficit_count + deficit_debt

            best = max(candidates, key=score)
            assignment[grp["id"]] = best
            current_count[best] += 1
            current_debt[best] += grp["debt"]

        if unassignable:
            raise KeyError(
                "STEP::تعذر إيجاد محصل بديل لبعض العملاء (رقم الهوية)::MISSING::"
                + "|".join(unassignable[:25])
            )

        # ------------------------------------------------------
        # 🔧 تحسين محلي: تقليل الفرق في العدد والمديونية أكتر من التوزيع
        # الأولي (Greedy) عن طريق Move + Swap بين المحصلين
        # ------------------------------------------------------
        refine_assignment(
            groups, assignment, current_count, current_debt,
            target_count, target_debt, collectors
        )

        df["المحصل الجديد"] = df["رقم الهوية"].map(assignment)

        # جدول مقارنة قبل / بعد لكل محصل
        summary_rows = []
        for c in collectors:
            summary_rows.append({
                "المحصل": c,
                "عدد العملاء (قديم)": target_count.get(c, 0),
                "عدد العملاء (جديد)": current_count.get(c, 0),
                "فرق العدد": current_count.get(c, 0) - target_count.get(c, 0),
                "متبقي المديونية (قديم)": target_debt.get(c, 0.0),
                "متبقي المديونية (جديد)": current_debt.get(c, 0.0),
                "فرق المديونية": current_debt.get(c, 0.0) - target_debt.get(c, 0.0),
            })
        summary_df = pd.DataFrame(summary_rows)

        return df.reset_index(drop=True), summary_df

    if rotation_file:
        try:
            file_bytes = rotation_file.getvalue()
            result_df, summary_df = rotate_portfolio(file_bytes)

            # ------------------------------------------------------
            # تحقق نهائي (Sanity check) من الشرطين الأساسيين
            # ------------------------------------------------------
            same_collector_violations = int(
                (result_df["المحصل الجديد"] == result_df["اسم المحصل القديم"]).sum()
            )
            split_id_violations = int(
                result_df.groupby("رقم الهوية")["المحصل الجديد"].nunique().gt(1).sum()
            )

            total_clients = result_df["رقم الهوية"].nunique()

            # ==========================================
            # 🔢 بطاقات KPI
            # ==========================================
            st.markdown(f"""
            <div class="kpi-grid-3">
                <div class="kpi-card ok">
                    <div class="kpi-icon">👥</div>
                    <div class="kpi-label">عدد العملاء اللي اتدوروا</div>
                    <div class="kpi-value">{total_clients:,}</div>
                    <div class="kpi-sub">على {len(result_df):,} حساب</div>
                </div>
                <div class="kpi-card {'ok' if same_collector_violations == 0 else ''}">
                    <div class="kpi-icon">{'✅' if same_collector_violations == 0 else '⚠️'}</div>
                    <div class="kpi-label">صفوف احتفظت بنفس المحصل</div>
                    <div class="kpi-value">{same_collector_violations:,}</div>
                    <div class="kpi-sub">لازم تكون صفر دايمًا</div>
                </div>
                <div class="kpi-card {'ok' if split_id_violations == 0 else ''}">
                    <div class="kpi-icon">{'✅' if split_id_violations == 0 else '⚠️'}</div>
                    <div class="kpi-label">هويات اتوزعت على أكتر من محصل</div>
                    <div class="kpi-value">{split_id_violations:,}</div>
                    <div class="kpi-sub">لازم تكون صفر دايمًا</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            if same_collector_violations == 0 and split_id_violations == 0:
                st.markdown(
                    '<div class="success-box">✅ التوزيع الجديد يحقق الشرطين بالكامل: '
                    'مفيش أي عميل احتفظ بمحصله القديم، ومفيش أي هوية اتوزعت على أكتر من محصل. '
                    'وتم كمان تشغيل مرحلة تحسين محلي لتقليل الفروق في العدد والمديونية قدر الإمكان.</div>',
                    unsafe_allow_html=True
                )

            # ==========================================
            # تحميل النتيجة
            # ==========================================
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="التدوير")
                summary_df.to_excel(writer, index=False, sheet_name="مقارنة قبل وبعد")
            output.seek(0)

            st.download_button(
                "📥 تحميل ملف التدوير (مع جدول المقارنة)",
                data=output,
                file_name="التدوير.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # ==========================================
            # جدول المقارنة + البيانات
            # ==========================================
            st.markdown('<div class="section-title">📊 مقارنة قبل وبعد لكل محصل</div>', unsafe_allow_html=True)

            def style_summary(d):
                fmt = {
                    "عدد العملاء (قديم)": "{:,.0f}", "عدد العملاء (جديد)": "{:,.0f}", "فرق العدد": "{:+,.0f}",
                    "متبقي المديونية (قديم)": "{:,.0f}", "متبقي المديونية (جديد)": "{:,.0f}", "فرق المديونية": "{:+,.0f}",
                }
                return d.style.format(fmt)

            tab_summary, tab_data = st.tabs(["📄 جدول المقارنة", "🗂️ البيانات الكاملة بعد التدوير"])

            with tab_summary:
                st.dataframe(style_summary(summary_df), use_container_width=True, hide_index=True)

            with tab_data:
                st.dataframe(result_df, use_container_width=True, hide_index=True)

            # ==========================================
            # الرسوم البيانية
            # ==========================================
            st.markdown('<div class="section-title">📈 الرسوم البيانية</div>', unsafe_allow_html=True)

            BLUE = "#155a8a"
            GOLD = "#C9A227"

            def get_theme_text_color():
                try:
                    base = st.get_option("theme.base")
                except Exception:
                    base = None
                return "#FFFFFF" if base == "dark" else "#111827"

            AXIS_TEXT_COLOR = get_theme_text_color()
            GRID_COLOR = "#374151" if AXIS_TEXT_COLOR == "#FFFFFF" else "#f1f5f3"
            LABEL_FONT = dict(size=13, family="Tajawal", color=AXIS_TEXT_COLOR)

            def style_fig(fig, angle=-20):
                fig.update_xaxes(tickfont=dict(size=13, family="Tajawal", color=AXIS_TEXT_COLOR))
                fig.update_yaxes(tickfont=dict(size=12, family="Tajawal", color=AXIS_TEXT_COLOR), gridcolor=GRID_COLOR)
                fig.update_layout(
                    height=400, xaxis_tickangle=angle, margin=dict(t=20, b=10, l=10, r=10),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Tajawal"),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                return fig

            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.markdown('<div class="chart-card"><div class="chart-card-title">عدد العملاء: قديم مقابل جديد لكل محصل</div>', unsafe_allow_html=True)
                count_melt = summary_df.melt(
                    id_vars="المحصل",
                    value_vars=["عدد العملاء (قديم)", "عدد العملاء (جديد)"],
                    var_name="النوع", value_name="عدد العملاء"
                )
                fig1 = px.bar(
                    count_melt, x="المحصل", y="عدد العملاء", color="النوع",
                    barmode="group", text="عدد العملاء",
                    color_discrete_map={"عدد العملاء (قديم)": GOLD, "عدد العملاء (جديد)": BLUE},
                    template="plotly_white"
                )
                fig1.update_traces(texttemplate="<b>%{text:,.0f}</b>", textposition="outside", textfont=LABEL_FONT)
                st.plotly_chart(style_fig(fig1), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

            with chart_col2:
                st.markdown('<div class="chart-card"><div class="chart-card-title">متبقي المديونية: قديم مقابل جديد لكل محصل</div>', unsafe_allow_html=True)
                debt_melt = summary_df.melt(
                    id_vars="المحصل",
                    value_vars=["متبقي المديونية (قديم)", "متبقي المديونية (جديد)"],
                    var_name="النوع", value_name="متبقي المديونية"
                )
                fig2 = px.bar(
                    debt_melt, x="المحصل", y="متبقي المديونية", color="النوع",
                    barmode="group", text="متبقي المديونية",
                    color_discrete_map={"متبقي المديونية (قديم)": GOLD, "متبقي المديونية (جديد)": BLUE},
                    template="plotly_white"
                )
                fig2.update_traces(texttemplate="<b>%{text:,.0f}</b>", textposition="outside", textfont=LABEL_FONT)
                st.plotly_chart(style_fig(fig2), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

        except KeyError as e:
            show_error(e)
        except Exception as e:
            show_error(e)
    else:
        st.markdown('<div class="empty-state">⬆️ ارفع ملف المحفظة عشان يبدأ التدوير</div>', unsafe_allow_html=True)

elif page == "وعود لا يوجد لها تاريخ الوعد":
    st.subheader("📅 وعود بدون تاريخ")
